"""
merge_reports.py — 合并多份 session_report.xlsx，生成总体报告

用法:
    python merge_reports.py path1/session_report.xlsx path2/session_report.xlsx --out merged_report/
    python merge_reports.py reports/*.xlsx --out /tmp/merged

输出（写入 --out 目录）:
    merged_session_report.xlsx   — 合并详情（新增 source 列）+ 分布统计
    merged_session_report.html   — 总体 HTML 报告（复用 analyze_sessions 模板）
    merged_session_report.md     — 总体 Markdown 报告
"""

import argparse
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd


# ---------------------------------------------------------------------------
# 从 analyze_sessions 复用核心逻辑
# ---------------------------------------------------------------------------

def _import_analyze():
    """动态导入同目录的 analyze_sessions 模块。"""
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "analyze_sessions",
        Path(__file__).parent / "analyze_sessions.py",
    )
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


# ---------------------------------------------------------------------------
# Excel 读取 + 合并
# ---------------------------------------------------------------------------

# Session详情 sheet 里的中文列名 → 内部字段名（与 analyze_sessions._DETAIL_COLS 对应）
_LABEL_TO_KEY = {
    "Q1首问":            "q1",
    "Session":           "session",
    "开始时间":           "start_time",
    "结束时间":           "end_time",
    "持续时长(s)":        "duration_s",
    "请求次数":           "api_call_count",
    "API错误次数":        "api_errors",
    "用户轮次":           "user_turns",
    "消息总数":           "total_messages",
    "tool_use次数":       "tool_use_count",
    "tool_result次数":    "tool_result_count",
    "工具成功次数":        "tool_success",
    "失败(is_error标记)": "tool_fail_flag",
    "失败(错误关键字)":    "tool_fail_keyword",
    "失败合计":           "tool_fail_total",
    "工具成功率(%)":      "tool_success_rate",
    "模型":              "model",
    "工具调用详情":        "tool_use_detail",
    "工具成功详情":        "tool_success_detail",
    "使用的技能":         "skills_used",
    "任务完成":           "completed",
    "错误备注":           "completed_note",
}

_TOOLS_SUCCESS_LABELS = {"工具成功详情", "工具成功分布"}
_TOOLS_FAIL_LABELS = {"工具失败详情", "工具失败分布"}

_INT_COLS   = {"api_call_count", "api_errors", "user_turns", "total_messages",
               "tool_use_count", "tool_result_count", "tool_success",
               "tool_fail_flag", "tool_fail_keyword", "tool_fail_total"}
_FLOAT_COLS = {"duration_s", "tool_success_rate"}


def _parse_tool_dict(s: str) -> dict:
    """将 'exec:1138, read:130' 反解析回 {'exec': 1138, 'read': 130}。"""
    if not isinstance(s, str) or not s.strip():
        return {}
    result = {}
    for part in s.split(","):
        part = part.strip()
        if ":" not in part:
            continue
        name, _, count = part.rpartition(":")
        name = name.strip()
        try:
            result[name] = int(count.strip())
        except ValueError:
            pass
    return result


def _parse_skills_str(s: str) -> dict:
    """将 'skillA, skillB' 反解析回 {'skillA': 1, 'skillB': 1}。"""
    if not isinstance(s, str) or not s.strip():
        return {}
    return {name.strip(): 1 for name in s.split(",") if name.strip()}


def _coerce(key: str, val: Any) -> Any:
    if pd.isna(val) if not isinstance(val, (list, dict)) else False:
        return None
    if key == "completed":
        text = str(val).strip()
        return 0 if text == "0" else text
    if key == "tool_use_detail":
        if isinstance(val, dict):
            return val
        return _parse_tool_dict(str(val)) if val else {}
    if key == "tool_success_detail":
        if isinstance(val, dict):
            return val
        return _parse_tool_dict(str(val)) if val else {}
    if key == "skills_used":
        if isinstance(val, dict):
            return val
        return _parse_skills_str(str(val)) if val else {}
    if key in _INT_COLS:
        try:
            return int(float(val))
        except (TypeError, ValueError):
            return 0
    if key in _FLOAT_COLS:
        try:
            v = float(val)
            # tool_success_rate 在 Excel 里可能已被存为小数（0.95 表示 95%）
            if key == "tool_success_rate" and 0.0 < v <= 1.0:
                v = round(v * 100, 1)
            return v
        except (TypeError, ValueError):
            return None
    return val


def load_xlsx(path: Path) -> List[Dict]:
    """读取一份 session_report.xlsx 的 Session详情 sheet，返回 sessions 列表。"""
    try:
        df = pd.read_excel(path, sheet_name="Session详情", dtype=str, engine="openpyxl")
    except Exception as e:
        print(f"[warn] 无法读取 {path}: {e}", file=sys.stderr)
        return []

    sessions = []
    for _, row in df.iterrows():
        s: Dict = {}
        for col_label, val in row.items():
            key = _LABEL_TO_KEY.get(str(col_label))
            if key is None:
                continue
            s[key] = _coerce(key, val)

        # 补全可能缺失的字段
        s.setdefault("q1", "")
        s.setdefault("session", "")
        s.setdefault("model", "")
        s.setdefault("completed", 0)
        s.setdefault("completed_note", "")
        s.setdefault("tool_use_detail", {})
        s.setdefault("tool_success_detail", {})
        s.setdefault("tool_fail_detail", {})
        s.setdefault("skills_used", {})

        sessions.append(s)
    return sessions


# ---------------------------------------------------------------------------
# 合并 Excel（含 source 列）
# ---------------------------------------------------------------------------

def write_merged_excel(
    all_sessions: List[Dict],
    source_map: Dict[str, str],   # session_id → source 路径标签
    stats: Dict,
    path: Path,
    az: Any,                       # analyze_sessions 模块
) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    _DETAIL_COLS_MERGE = [("source", "来源文件")] + list(az._DETAIL_COLS)
    header_labels = [label for _, label in _DETAIL_COLS_MERGE]

    rows = []
    for s in all_sessions:
        row: Dict = {}
        for key, label in _DETAIL_COLS_MERGE:
            if key == "source":
                row[label] = az._sanitize_cell(source_map.get(s.get("session", ""), ""))
            elif key == "tool_use_detail":
                row[label] = az._sanitize_cell(az._fmt_tool_dict(s.get(key)))
            else:
                row[label] = az._sanitize_cell(s.get(key))
        rows.append(row)

    col_widths: Dict[str, int] = {}
    for label in header_labels:
        if label == "Q1首问":
            col_widths[label] = 50
            continue
        if label == "来源文件":
            col_widths[label] = 30
            continue
        max_len = len(label)
        for row in rows:
            val = row.get(label)
            if val is not None:
                max_len = max(max_len, len(str(val)))
        col_widths[label] = min(max_len + 2, 40)

    df_detail = pd.DataFrame(rows)

    turns_vals = stats["turns_vals"]
    msg_vals   = stats["msg_vals"]
    api_vals   = stats["api_vals"]
    tu_vals    = stats["tu_vals"]
    rate_vals  = stats["rate_vals"]
    dur_vals   = stats["dur_vals"]

    dist_sections = [
        ("对话轮次分布",
         az._bucket_df(pd, turns_vals,
                       [(1,1,2),(2,2,4),(4,4,8),(8,8,16),(16,16,10**9)], "轮")),
        ("消息总数分布",
         az._bucket_df(pd, msg_vals,
                       [("1-2",0,3),("3-5",3,6),("6-10",6,11),
                        ("11-20",11,21),("21-50",21,51),(">50",51,10**9)], "条")),
        ("API Call次数分布",
         az._bucket_df(pd, api_vals,
                       [("1次",1,2),("2-3次",2,4),("4-10次",4,11),
                        ("11-30次",11,31),(">30次",31,10**9)])),
        ("tool_use次数分布（有工具session）",
         az._bucket_df(pd, tu_vals,
                       [("1-5次",1,6),("6-15次",6,16),("16-30次",16,31),
                        ("31-50次",31,51),(">50次",51,10**9)]) if tu_vals else pd.DataFrame()),
        ("工具成功率分布",
         az._bucket_df(pd, rate_vals,
                       [("0-50%",0,50),("50-80%",50,80),("80-95%",80,95),
                        ("95-99%",95,99),("100%",100,101)]) if rate_vals else pd.DataFrame()),
        ("耗时分布（多轮session）",
         az._bucket_df(pd, dur_vals,
                       [("<1min",0,60),("1-5min",60,300),("5-15min",300,900),
                        ("15-30min",900,1800),(">30min",1800,10**9)]) if dur_vals else pd.DataFrame()),
    ]

    with pd.ExcelWriter(path, engine="openpyxl") as ew:
        # Sheet1: 合并详情
        df_detail.to_excel(ew, sheet_name="Session详情", index=False)
        ws1 = ew.sheets["Session详情"]
        ws1.freeze_panes = "C2"  # 冻结前两列（来源+Session）和首行

        hdr_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
        hdr_font = Font(bold=True, color="FFFFFF", size=10)
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        wrap_top_align = Alignment(wrap_text=True, vertical="top")
        for cell in ws1[1]:
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = hdr_align
        ws1.row_dimensions[1].height = 30

        alt_fill = PatternFill(fill_type="solid", fgColor="EBF3FB")
        for row_idx, row_cells in enumerate(
            ws1.iter_rows(min_row=2, max_row=ws1.max_row),
            start=2,
        ):
            if row_idx % 2 != 0:
                continue
            for cell in row_cells:
                if cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF", "00FFFFFF"):
                    cell.fill = alt_fill

        # 成功率列格式
        for ci, (_, lbl) in enumerate(_DETAIL_COLS_MERGE, 1):
            if lbl == "工具成功率(%)":
                for (cell,) in ws1.iter_rows(
                    min_row=2,
                    max_row=ws1.max_row,
                    min_col=ci,
                    max_col=ci,
                ):
                    if cell.value is not None:
                        try:
                            cell.value = float(cell.value) / 100
                            cell.number_format = "0.0%"
                        except (TypeError, ValueError):
                            pass
            if lbl in ("Q1首问", "来源文件"):
                for (cell,) in ws1.iter_rows(
                    min_row=2,
                    max_row=ws1.max_row,
                    min_col=ci,
                    max_col=ci,
                ):
                    cell.alignment = wrap_top_align

        # 列宽
        for col_idx, label in enumerate(header_labels, start=1):
            col_letter = get_column_letter(col_idx)
            ws1.column_dimensions[col_letter].width = col_widths[label]

        # Sheet2: 分布统计
        ws2 = ew.book.create_sheet("分布统计")
        title_font  = Font(bold=True, size=11, color="1F4E79")
        subhdr_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
        subhdr_font = Font(bold=True, size=10)

        cur_row = 1
        for section_title, df_s in dist_sections:
            if df_s.empty:
                continue
            ws2.cell(cur_row, 1, section_title).font = title_font
            cur_row += 1
            for ci, col_name in enumerate(df_s.columns, 1):
                hc = ws2.cell(cur_row, ci, col_name)
                hc.fill = subhdr_fill
                hc.font = subhdr_font
                hc.alignment = Alignment(horizontal="center")
            cur_row += 1
            for row in df_s.itertuples(index=False, name=None):
                for ci, val in enumerate(row, 1):
                    ws2.cell(cur_row, ci, val)
                cur_row += 1
            cur_row += 2

        for col in ws2.columns:
            ws2.column_dimensions[get_column_letter(col[0].column)].width = 32

        # Sheet3: 来源汇总
        ws3 = ew.book.create_sheet("来源汇总")
        from collections import Counter
        src_counter = Counter(source_map[s.get("session","")] for s in all_sessions)
        ws3.cell(1, 1, "来源文件").font = Font(bold=True)
        ws3.cell(1, 2, "session 数").font = Font(bold=True)
        for ri, (src, cnt) in enumerate(sorted(src_counter.items()), start=2):
            ws3.cell(ri, 1, src)
            ws3.cell(ri, 2, cnt)
        ws3.column_dimensions["A"].width = 50
        ws3.column_dimensions["B"].width = 12


# ---------------------------------------------------------------------------
# 主入口
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="合并多份 session_report.xlsx，生成总体报告",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "示例:\n"
            "  python merge_reports.py path1/session_report.xlsx path2/session_report.xlsx --out merged/\n"
            "  python merge_reports.py reports/*.xlsx --out /tmp/merged"
        ),
    )
    parser.add_argument(
        "xlsx_files", nargs="+", metavar="XLSX",
        help="一个或多个 session_report.xlsx 路径",
    )
    parser.add_argument(
        "--out", "-o", default="merged_report", metavar="OUTPUT_DIR",
        help="输出目录，默认 merged_report/",
    )
    args = parser.parse_args()

    xlsx_paths = [Path(p).resolve() for p in args.xlsx_files]
    missing = [p for p in xlsx_paths if not p.is_file()]
    if missing:
        for p in missing:
            print(f"[error] 文件不存在: {p}", file=sys.stderr)
        sys.exit(1)

    out_dir = Path(args.out).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    # 导入 analyze_sessions 以复用函数
    print("[info] 导入 analyze_sessions ...", file=sys.stderr)
    az = _import_analyze()

    # 读取并合并所有 xlsx
    all_sessions: List[Dict] = []
    source_map: Dict[str, str] = {}   # session_id → 来源标签

    for xlsx_path in xlsx_paths:
        label = str(xlsx_path)   # 完整路径作为来源标签
        print(f"[info] 读取: {xlsx_path}", file=sys.stderr)
        sessions = load_xlsx(xlsx_path)
        print(f"       → {len(sessions)} 条 session", file=sys.stderr)
        for s in sessions:
            sid = s.get("session", "")
            # 若多个文件含同名 session，加上来源区分
            unique_key = f"{label}::{sid}"
            source_map[sid] = label
            # 用 unique_key 作为内部 session id，避免覆盖统计
            s["_source"] = label
        all_sessions.extend(sessions)

    if not all_sessions:
        print("[error] 没有可用的 session 数据", file=sys.stderr)
        sys.exit(1)

    print(f"[info] 合并后共 {len(all_sessions)} 条 session", file=sys.stderr)

    # 重新计算汇总统计
    stats = az.compute_stats(all_sessions)
    ctx   = az.build_context(all_sessions, stats)
    # 在标题里标注来源
    ctx["generated_at"] += f"  (合并自 {len(xlsx_paths)} 份报告)"
    ctx["total_sources"] = len(xlsx_paths)

    # 写 Excel
    xlsx_out = out_dir / "merged_session_report.xlsx"
    write_merged_excel(all_sessions, source_map, stats, xlsx_out, az)
    print(f"[done] Excel    → {xlsx_out}", file=sys.stderr)

    # 写 HTML
    html_out = out_dir / "merged_session_report.html"
    az.render_report("report.html.j2", ctx, html_out)
    print(f"[done] HTML     → {html_out}", file=sys.stderr)

    # 写 Markdown
    md_out = out_dir / "merged_session_report.md"
    az.render_report("report.md.j2", ctx, md_out)
    print(f"[done] Markdown → {md_out}", file=sys.stderr)


if __name__ == "__main__":
    main()

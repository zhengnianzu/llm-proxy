"""
analyze_sessions.py — 分析 session 格式对话日志

用法:
    python analyze_sessions.py --dir <目录A> --dir <目录B> --out output

输出（单目录默认写入 <目录>/stat/；多目录写入 --out/<目录名>/）:
    session_report.xlsx   — 每条 session 详情 + 分布统计（两个 sheet）
    session_report.md     — 汇总概览报告
"""

import argparse
import json
import logging
import os
import re
import sys
import time
from collections import Counter
from concurrent.futures import ProcessPoolExecutor, as_completed
from tqdm import tqdm
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import yaml

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)


def format_stage_seconds(seconds: float) -> str:
    return f"{seconds:.2f}s"

# ---------------------------------------------------------------------------
# 时间戳解析
# ---------------------------------------------------------------------------

FNAME_TS_RE = re.compile(r"^(\d{4}-\d{2}-\d{2})_(\d{2})-(\d{2})-(\d{2})_\d{3}$")


def parse_folder_ts(name: str) -> Optional[datetime]:
    m = FNAME_TS_RE.match(name)
    if not m:
        return None
    try:
        return datetime.strptime(
            f"{m.group(1)} {m.group(2)}:{m.group(3)}:{m.group(4)}",
            "%Y-%m-%d %H:%M:%S",
        )
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# 工具错误关键字
# ---------------------------------------------------------------------------

_ERROR_PATTERNS: List[re.Pattern] = [
    re.compile(r"Traceback \(most recent call last\)", re.I),
    re.compile(
        r"\b(SyntaxError|NameError|TypeError|ValueError|AttributeError|"
        r"ImportError|ModuleNotFoundError|RuntimeError|KeyError|IndexError|"
        r"FileNotFoundError|PermissionError|OSError|IOError|ZeroDivisionError|"
        r"RecursionError|MemoryError)\s*:", re.I
    ),
    re.compile(r"permission denied", re.I),
    re.compile(r"operation not permitted", re.I),
    re.compile(r"access denied", re.I),
    re.compile(r"cannot execute", re.I),
    re.compile(r"no such file or directory", re.I),
    re.compile(r"file not found", re.I),
    re.compile(r"\btimed?\s*out\b", re.I),
    re.compile(r"\bkilled\b", re.I),
    re.compile(r"segmentation fault", re.I),
    re.compile(r"command not found", re.I),
    re.compile(r"\berror\b", re.I),
    re.compile(r"\bfailed\b", re.I),
    re.compile(r"\bexception\b", re.I),
    re.compile(r"\bfailure\b", re.I),
]


def _has_error_keywords(text: str) -> bool:
    return any(p.search(text) for p in _ERROR_PATTERNS)


def _collect_text(content: Any) -> str:
    if isinstance(content, str):
        return content
    if isinstance(content, list):
        parts = []
        for blk in content:
            if not isinstance(blk, dict):
                continue
            if blk.get("text"):
                parts.append(str(blk["text"]))
            inner = blk.get("content", "")
            if isinstance(inner, str):
                parts.append(inner)
            elif isinstance(inner, list):
                for sub in inner:
                    if isinstance(sub, dict) and sub.get("text"):
                        parts.append(str(sub["text"]))
        return "\n".join(parts)
    return ""


# ---------------------------------------------------------------------------
# 内容遍历
# ---------------------------------------------------------------------------

def iter_blocks(content: Any):
    if isinstance(content, str):
        yield {"type": "text", "text": content}
    elif isinstance(content, list):
        for b in content:
            if isinstance(b, dict):
                yield b


def count_user_turns(messages: List[dict]) -> int:
    count = 0
    for m in messages:
        if m.get("role") != "user":
            continue
        blocks = list(iter_blocks(m.get("content", [])))
        if any(b.get("type") != "tool_result" for b in blocks):
            count += 1
    return count


def count_tool_use(messages: List[dict], resp_content: Optional[List]) -> int:
    count = 0
    for m in messages:
        if m.get("role") == "assistant":
            for b in iter_blocks(m.get("content", [])):
                if b.get("type") == "tool_use":
                    count += 1
    if isinstance(resp_content, list):
        for b in resp_content:
            if isinstance(b, dict) and b.get("type") == "tool_use":
                count += 1
    return count


def analyze_tool_results(messages: List[dict]) -> Dict[str, int]:
    total = success = fail_flag = fail_kw = 0
    for m in messages:
        if m.get("role") != "user":
            continue
        for blk in iter_blocks(m.get("content", [])):
            if blk.get("type") != "tool_result":
                continue
            total += 1
            if blk.get("is_error"):
                fail_flag += 1
            else:
                text = _collect_text(blk.get("content", ""))
                if _has_error_keywords(text):
                    fail_kw += 1
                else:
                    success += 1
    fail_total = fail_flag + fail_kw
    return {
        "total": total,
        "success": total - fail_total,
        "fail_flag": fail_flag,
        "fail_kw": fail_kw,
        "fail_total": fail_total,
    }


def analyze_tools_detail(full_messages: List[dict]) -> Dict[str, Dict[str, int]]:
    """统计每个工具的调用次数和成功/失败次数。

    返回 {"use": {tool_name: count}, "success": {...}, "fail": {...}}
    """
    id_to_name: Dict[str, str] = {}
    use_counts: Counter = Counter()
    success_counts: Counter = Counter()
    fail_counts: Counter = Counter()

    for m in full_messages:
        for blk in iter_blocks(m.get("content", [])):
            if blk.get("type") == "tool_use":
                tool_id = blk.get("id", "")
                name = blk.get("name", "unknown")
                if tool_id:
                    id_to_name[tool_id] = name
                use_counts[name] += 1

    for m in full_messages:
        if m.get("role") != "user":
            continue
        for blk in iter_blocks(m.get("content", [])):
            if blk.get("type") != "tool_result":
                continue
            name = id_to_name.get(blk.get("tool_use_id", ""), "unknown")
            if blk.get("is_error"):
                fail_counts[name] += 1
            else:
                text = _collect_text(blk.get("content", ""))
                if _has_error_keywords(text):
                    fail_counts[name] += 1
                else:
                    success_counts[name] += 1

    return {
        "use":     dict(use_counts),
        "success": dict(success_counts),
        "fail":    dict(fail_counts),
    }


_SKILL_PATH_RE = re.compile(r"(?:^|/)\.openclaw/skills/([^/]+)/SKILL\.md")


def analyze_skills(full_messages: List[dict]) -> Dict[str, int]:
    """统计使用的自定义技能。

    识别条件：工具名为 'read'，且 input.file_path 匹配 .openclaw/skills/xxx/SKILL.md。
    技能名取 xxx 部分。

    返回 {skill_name: count}
    """
    skill_counts: Counter = Counter()
    for m in full_messages:
        for blk in iter_blocks(m.get("content", [])):
            if blk.get("type") != "tool_use" or blk.get("name") != "read":
                continue
            inp = blk.get("input") or {}
            if isinstance(inp, str):
                try:
                    import json as _json
                    inp = _json.loads(inp)
                except Exception:
                    inp = {}
            if not isinstance(inp, dict):
                inp = {}
            file_path = inp.get("file_path") or inp.get("path") or ""
            if "SKILL.md" not in file_path:
                continue
            match = _SKILL_PATH_RE.search(file_path)
            if match:
                skill_counts[match.group(1)] += 1
    return dict(skill_counts)


# ---------------------------------------------------------------------------
# Q1 提取
# ---------------------------------------------------------------------------

def get_q1(messages: List[dict]) -> str:
    for msg in messages:
        if msg.get("role") != "user":
            continue
        c = msg.get("content", "")
        if isinstance(c, str):
            text = c
        elif isinstance(c, list):
            parts = [b.get("text") or b.get("id") or "" for b in c if isinstance(b, dict)]
            text = "\n".join(p for p in parts if p)
        else:
            text = str(c)
        while True:
            prev = text
            text = re.sub(r"^\s*\[[^\]]*\]\s*", "", text)
            text = re.sub(
                r"^Sender\s*(?:\([^)]*\))?:\s*```json\s*\{[\s\S]*?\}\s*```\s*",
                "", text, flags=re.IGNORECASE,
            )
            text = re.sub(r"^Sender\s*(?:\([^)]*\))?:[^\n]*\n?", "", text, flags=re.IGNORECASE)
            if text == prev:
                break
        return text.strip()
    return ""


# ---------------------------------------------------------------------------
# 质量评估
# ---------------------------------------------------------------------------

# 错误代码 → 中文说明
QUALITY_ERRORS: Dict[str, str] = {
    "E001": "乱码(行均字符过少)",
    "E002": "200空响应",
    "E003": "工具调用过少(<3次)",
}


def _is_garbled(text: str, min_lines: int = 10, max_avg_chars: float = 5.0) -> bool:
    """判断文本是否疑似乱码：非空行数 >= min_lines 且平均每行字符数 < max_avg_chars。"""
    lines = [l for l in text.splitlines() if l.strip()]
    if len(lines) < min_lines:
        return False
    return (sum(len(l) for l in lines) / len(lines)) < max_avg_chars


def _scan_garbled(data: dict) -> bool:
    """遍历单个请求的 messages content、thinking 块及 response content，检测乱码。"""
    # messages
    for msg in (data.get("messages") or []):
        content = msg.get("content", [])
        # 普通文本内容
        text = _collect_text(content)
        if text and _is_garbled(text):
            return True
        # thinking / reasoning 块
        if isinstance(content, list):
            for blk in content:
                if not isinstance(blk, dict):
                    continue
                thinking = blk.get("thinking") or blk.get("reasoning_content") or ""
                if isinstance(thinking, str) and _is_garbled(thinking):
                    return True
    # response content
    resp_content = (data.get("response") or {}).get("content")
    if resp_content:
        text = _collect_text(resp_content)
        if text and _is_garbled(text):
            return True
    return False


def check_quality(best_data: dict, tool_use_cnt: int) -> List[str]:
    """
    对单个 session 的最佳快照做质量检查。
    返回触发的错误代码列表（空列表 = 无问题）。
    """
    errors: List[str] = []

    # E001: 乱码
    if _scan_garbled(best_data):
        errors.append("E001")

    # E002: 请求返回 200 但响应内容为空
    resp = best_data.get("response") or {}
    status = resp.get("status_code")
    content = resp.get("content")
    if status == 200 and not content:
        errors.append("E002")

    # E003: 工具调用次数 < 3
    if tool_use_cnt < 3:
        errors.append("E003")

    return errors


def fmt_quality(error_codes: List[str]) -> tuple:
    """返回 (codes, note) 二元组。无错误时 codes=0, note=''。"""
    if not error_codes:
        return 0, ""
    codes = ",".join(error_codes)
    note  = "; ".join(QUALITY_ERRORS[c] for c in error_codes if c in QUALITY_ERRORS)
    return codes, note


def _extract_first_user_text_from_content(content: Any) -> str:
    if isinstance(content, str):
        text = content
    elif isinstance(content, list):
        parts = [b.get("text") or b.get("id") or "" for b in content if isinstance(b, dict)]
        text = "\n".join(p for p in parts if p)
    else:
        text = str(content)

    while True:
        prev = text
        text = re.sub(r"^\s*\[[^\]]*\]\s*", "", text)
        text = re.sub(
            r"^Sender\s*(?:\([^)]*\))?:\s*```json\s*\{[\s\S]*?\}\s*```\s*",
            "", text, flags=re.IGNORECASE,
        )
        text = re.sub(r"^Sender\s*(?:\([^)]*\))?:[^\n]*\n?", "", text, flags=re.IGNORECASE)
        if text == prev:
            break
    return text.strip()


def _new_best_data_stats() -> Dict[str, Any]:
    return {
        "q1": "",
        "user_turns": 0,
        "tool_use_count": 0,
        "tool_result_count": 0,
        "tool_success": 0,
        "tool_fail_flag": 0,
        "tool_fail_keyword": 0,
        "tool_use_detail": Counter(),
        "tool_success_detail": Counter(),
        "tool_fail_detail": Counter(),
        "skills_used": Counter(),
        "id_to_name": {},
        "has_garbled": False,
    }


def _mark_garbled_from_content(content: Any, stats: Dict[str, Any]) -> None:
    text = _collect_text(content)
    if text and _is_garbled(text):
        stats["has_garbled"] = True
    if isinstance(content, list):
        for blk in content:
            if not isinstance(blk, dict):
                continue
            thinking = blk.get("thinking") or blk.get("reasoning_content") or ""
            if isinstance(thinking, str) and _is_garbled(thinking):
                stats["has_garbled"] = True


def _record_skill_use(name: str, blk: dict, stats: Dict[str, Any]) -> None:
    if name != "read":
        return
    inp = blk.get("input") or {}
    if isinstance(inp, str):
        try:
            inp = json.loads(inp)
        except Exception:
            inp = {}
    if not isinstance(inp, dict):
        inp = {}
    file_path = inp.get("file_path") or inp.get("path") or ""
    if "SKILL.md" not in file_path:
        return
    match = _SKILL_PATH_RE.search(file_path)
    if match:
        stats["skills_used"][match.group(1)] += 1


def _record_tool_use(blk: dict, stats: Dict[str, Any]) -> None:
    tool_id = blk.get("id", "")
    name = blk.get("name", "unknown")
    if tool_id:
        stats["id_to_name"][tool_id] = name
    stats["tool_use_detail"][name] += 1
    stats["tool_use_count"] += 1
    _record_skill_use(name, blk, stats)


def _record_tool_result(blk: dict, stats: Dict[str, Any]) -> None:
    stats["tool_result_count"] += 1
    name = stats["id_to_name"].get(blk.get("tool_use_id", ""), "unknown")
    if blk.get("is_error"):
        stats["tool_fail_flag"] += 1
        stats["tool_fail_detail"][name] += 1
        return

    text = _collect_text(blk.get("content", ""))
    if _has_error_keywords(text):
        stats["tool_fail_keyword"] += 1
        stats["tool_fail_detail"][name] += 1
    else:
        stats["tool_success"] += 1
        stats["tool_success_detail"][name] += 1


def _analyze_user_message(content: Any, stats: Dict[str, Any]) -> None:
    has_non_tool_result = False
    for blk in iter_blocks(content):
        blk_type = blk.get("type")
        if blk_type != "tool_result":
            has_non_tool_result = True
        if blk_type == "tool_result":
            _record_tool_result(blk, stats)
    if has_non_tool_result:
        stats["user_turns"] += 1


def _analyze_assistant_message(content: Any, stats: Dict[str, Any]) -> None:
    for blk in iter_blocks(content):
        if blk.get("type") == "tool_use":
            _record_tool_use(blk, stats)


def _build_quality_errors(resp: dict, resp_content: Any, stats: Dict[str, Any]) -> List[str]:
    errors: List[str] = []
    if stats["has_garbled"]:
        errors.append("E001")
    if resp.get("status_code") == 200 and not resp_content:
        errors.append("E002")
    if stats["tool_use_count"] < 3:
        errors.append("E003")
    return errors


def analyze_best_data(best_data: dict) -> Dict[str, Any]:
    """在一次 pass 中完成单个 best_data 的主要统计。

    这里刻意保持与旧逻辑一致：
    - Q1 取第一条 user 消息文本
    - user_turns 仅在 user message 含非 tool_result block 时计数
    - tool_result 只从 user message 中统计
    - tool_use 同时统计 assistant messages 与最终 response.content
    """
    messages: List[dict] = best_data.get("messages") or []
    resp = best_data.get("response") or {}
    resp_content = resp.get("content")
    stats = _new_best_data_stats()

    for msg in messages:
        role = msg.get("role")
        content = msg.get("content", [])

        if role == "user" and not stats["q1"]:
            stats["q1"] = _extract_first_user_text_from_content(content)

        if role == "user":
            _analyze_user_message(content, stats)
        elif role == "assistant":
            _analyze_assistant_message(content, stats)

        # 质量检查依赖正文和 thinking/reasoning 内容，这里与主统计一并扫描。
        _mark_garbled_from_content(content, stats)

    if isinstance(resp_content, list):
        for blk in resp_content:
            if isinstance(blk, dict) and blk.get("type") == "tool_use":
                _record_tool_use(blk, stats)
        _mark_garbled_from_content(resp_content, stats)

    tool_result_count = stats["tool_result_count"]
    return {
        "q1": stats["q1"],
        "total_messages": len(messages) + (1 if resp_content else 0),
        "user_turns": stats["user_turns"],
        "tool_use_count": stats["tool_use_count"],
        "tool_result_count": tool_result_count,
        "tool_success": stats["tool_success"],
        "tool_fail_flag": stats["tool_fail_flag"],
        "tool_fail_keyword": stats["tool_fail_keyword"],
        "tool_fail_total": stats["tool_fail_flag"] + stats["tool_fail_keyword"],
        "tool_success_rate": (
            round(stats["tool_success"] / tool_result_count * 100, 1) if tool_result_count > 0 else None
        ),
        "tool_use_detail": dict(stats["tool_use_detail"]),
        "tool_success_detail": dict(stats["tool_success_detail"]),
        "tool_fail_detail": dict(stats["tool_fail_detail"]),
        "skills_used": dict(stats["skills_used"]),
        "quality_errors": _build_quality_errors(resp, resp_content, stats),
    }


# ---------------------------------------------------------------------------
# 单 session 分析
# ---------------------------------------------------------------------------

def analyze_session(folder: Path) -> Optional[Dict]:
    json_files = sorted(folder.glob("*.json"))
    if not json_files:
        return None

    start_ts = parse_folder_ts(folder.name)
    end_ts   = parse_folder_ts(json_files[-1].stem)
    duration_s: Optional[float] = None
    if start_ts and end_ts and end_ts >= start_ts:
        duration_s = (end_ts - start_ts).total_seconds()

    api_call_count = len(json_files)
    api_errors = 0

    try:
        best_data = json.loads(json_files[-1].read_text(encoding="utf-8"))
    except Exception:
        return None

    resp = best_data.get("response") or {}
    if isinstance(resp, dict) and isinstance(resp.get("status_code"), int):
        if resp["status_code"] >= 400:
            api_errors = 1

    analyzed = analyze_best_data(best_data)

    return {
        "session":            folder.name,
        "start_time":         start_ts.strftime("%Y-%m-%d %H:%M:%S") if start_ts else None,
        "end_time":           end_ts.strftime("%Y-%m-%d %H:%M:%S")   if end_ts   else None,
        "duration_s":         duration_s,
        "api_call_count":     api_call_count,
        "api_errors":         api_errors,
        "user_turns":         analyzed["user_turns"],
        "total_messages":     analyzed["total_messages"],
        "tool_use_count":     analyzed["tool_use_count"],
        "tool_result_count":  analyzed["tool_result_count"],
        "tool_success":       analyzed["tool_success"],
        "tool_fail_flag":     analyzed["tool_fail_flag"],
        "tool_fail_keyword":  analyzed["tool_fail_keyword"],
        "tool_fail_total":    analyzed["tool_fail_total"],
        "tool_success_rate":  analyzed["tool_success_rate"],
        "model":              best_data.get("model", ""),
        "q1":                 analyzed["q1"],
        "tool_use_detail":    analyzed["tool_use_detail"],
        "tool_success_detail": analyzed["tool_success_detail"],
        "tool_fail_detail":   analyzed["tool_fail_detail"],
        "skills_used":        analyzed["skills_used"],
        **dict(zip(("completed", "completed_note"),
                   fmt_quality(analyzed["quality_errors"]))),
    }


# ---------------------------------------------------------------------------
# 辅助
# ---------------------------------------------------------------------------

def fmt_duration(s: Optional[float]) -> str:
    if s is None:
        return "N/A"
    if s < 60:
        return f"{s:.0f}s"
    if s < 3600:
        return f"{s/60:.1f}min"
    return f"{s/3600:.1f}h"


def fmt_rate(r: Optional[float]) -> str:
    return f"{r:.1f}%" if r is not None else "-"


def pct(values: List[float], p: int) -> float:
    if not values:
        return 0.0
    sv = sorted(values)
    return sv[min(int(len(sv) * p / 100), len(sv) - 1)]


def _dist_rows(vals: List, buckets: List[Tuple], unit: str = "") -> List[Dict]:
    """将数值列表按桶分组，返回含标签/计数/占比/条形图的行列表，供模板直接渲染。"""
    total  = len(vals)
    counts = [sum(1 for v in vals if v is not None and lo <= v < hi) for _, lo, hi in buckets]
    max_c  = max(counts, default=1)
    return [
        {
            "label": f"{label}{unit}",
            "count": cnt,
            "pct":   round(cnt / total * 100, 1) if total else 0.0,
            "bar":   "█" * (max(1, round(cnt / max(max_c, 1) * 15)) if cnt else 0),
        }
        for (label, lo, hi), cnt in zip(buckets, counts)
    ]


def _extract_error_codes(completed) -> List[str]:
    """从 completed 字段值提取错误代码列表。"""
    if completed == 0:
        return []
    codes_part = str(completed).split(" ")[0]
    return [c.strip() for c in codes_part.split(",") if c.strip()]


def compute_stats(sessions: List[Dict]) -> Dict:
    """对 sessions 做一次性聚合，供 Excel 和模板渲染共用。"""
    turns_vals    = [s["user_turns"]        for s in sessions]
    msg_vals      = [s["total_messages"]    for s in sessions]
    api_vals      = [s["api_call_count"]    for s in sessions]
    with_tools    = [s for s in sessions if s["tool_use_count"] > 0]
    tu_vals       = [s["tool_use_count"]    for s in with_tools]
    rate_vals     = [s["tool_success_rate"] for s in sessions if s["tool_success_rate"] is not None]
    multi_timed   = [s for s in sessions if s["api_call_count"] > 1 and s["duration_s"] is not None]
    dur_vals      = [s["duration_s"]        for s in multi_timed]
    rate_sessions = [s for s in sessions if s["tool_result_count"] > 0]
    model_dist    = Counter(s["model"]      for s in sessions)

    # 全局工具调用统计
    global_use:     Counter = Counter()
    global_success: Counter = Counter()
    global_fail:    Counter = Counter()
    global_skills:  Counter = Counter()
    for s in sessions:
        global_use.update(s.get("tool_use_detail", {}))
        global_success.update(s.get("tool_success_detail", {}))
        global_fail.update(s.get("tool_fail_detail", {}))
        global_skills.update(s.get("skills_used", {}))

    return {
        "total":         len(sessions),
        "turns_vals":    turns_vals,
        "msg_vals":      msg_vals,
        "api_vals":      api_vals,
        "with_tools":    with_tools,
        "tu_vals":       tu_vals,
        "rate_vals":     rate_vals,
        "multi_timed":   multi_timed,
        "dur_vals":      dur_vals,
        "rate_sessions": rate_sessions,
        "model_dist":    model_dist,
        "multi_api":     sum(1 for v in api_vals if v > 1),
        "api_err":       sum(1 for s in sessions if s["api_errors"] > 0),
        "total_tu":      sum(s["tool_use_count"]    for s in sessions),
        "total_tr":      sum(s["tool_result_count"] for s in sessions),
        "total_succ":    sum(s["tool_success"]      for s in sessions),
        "total_ff":      sum(s["tool_fail_flag"]    for s in sessions),
        "total_fk":      sum(s["tool_fail_keyword"] for s in sessions),
        "total_ft":      sum(s["tool_fail_total"]   for s in sessions),
        "global_use":    global_use,
        "global_success": global_success,
        "global_fail":   global_fail,
        "global_skills": global_skills,
    }


def build_context(sessions: List[Dict], stats: Dict, top_n: int = 10) -> Dict:
    """将 sessions 和 stats 组装为 Jinja2 模板上下文，HTML 和 Markdown 模板共用。"""
    total         = stats["total"]
    turns_vals    = stats["turns_vals"]
    msg_vals      = stats["msg_vals"]
    api_vals      = stats["api_vals"]
    with_tools    = stats["with_tools"]
    tu_vals       = stats["tu_vals"]
    rate_sessions = stats["rate_sessions"]
    multi_timed   = stats["multi_timed"]
    dur_vals      = stats["dur_vals"]
    total_tr      = stats["total_tr"]
    single_count  = sum(1 for v in api_vals if v == 1)

    # 质量 / 完成状态统计
    ok_count    = sum(1 for s in sessions if s["completed"] == 0)
    fail_count  = total - ok_count
    err_counter: Counter = Counter()
    for s in sessions:
        for code in _extract_error_codes(s["completed"]):
            err_counter[code] += 1

    top_raw = sorted(
        [s for s in sessions if s.get("duration_s")],
        key=lambda x: -x["duration_s"],
    )[:top_n]

    def p(vals, q):  # shorthand
        return int(pct(vals, q))

    return {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "total":  total,
        "top_n":  top_n,

        "turns": {
            "avg":  f"{sum(turns_vals)/total:.2f}",
            "max":  max(turns_vals),
            "p50":  p(turns_vals, 50),
            "p90":  p(turns_vals, 90),
            "dist": _dist_rows(turns_vals, [
                ("1",1,2),("2-3",2,4),("4-7",4,8),("8-15",8,16),(">15",16,10**9)
            ], "轮"),
        },

        "messages": {
            "avg":  f"{sum(msg_vals)/total:.2f}",
            "max":  max(msg_vals),
            "min":  min(msg_vals),
            "p50":  p(msg_vals, 50),
            "p90":  p(msg_vals, 90),
            "dist": _dist_rows(msg_vals, [
                ("1-2",0,3),("3-5",3,6),("6-10",6,11),
                ("11-20",11,21),("21-50",21,51),(">50",51,10**9)
            ], "条"),
        },

        "api": {
            "avg":          f"{sum(api_vals)/total:.2f}",
            "max":          max(api_vals),
            "single_count": single_count,
            "single_pct":   f"{single_count/total*100:.1f}",
            "multi_count":  stats["multi_api"],
            "multi_pct":    f"{stats['multi_api']/total*100:.1f}",
            "err_count":    stats["api_err"],
            "err_pct":      f"{stats['api_err']/total*100:.1f}",
            "dist": _dist_rows(api_vals, [
                ("1次",1,2),("2-3次",2,4),("4-10次",4,11),("11-30次",11,31),(">30次",31,10**9)
            ]),
        },

        "tools": {
            "with_count":    len(with_tools),
            "with_pct":      f"{len(with_tools)/total*100:.1f}",
            "without_count": total - len(with_tools),
            "without_pct":   f"{(total-len(with_tools))/total*100:.1f}",
            "total_use":     stats["total_tu"],
            "total_result":  total_tr,
            "has_sessions":  bool(with_tools),
            "avg_use":       f"{sum(tu_vals)/len(tu_vals):.1f}" if tu_vals else "0",
            "max_use":       max(tu_vals) if tu_vals else 0,
            "p50_use":       p(tu_vals, 50) if tu_vals else 0,
            "p90_use":       p(tu_vals, 90) if tu_vals else 0,
            "has_results":   total_tr > 0,
            "total_succ":    stats["total_succ"],
            "total_ff":      stats["total_ff"],
            "total_fk":      stats["total_fk"],
            "total_ft":      stats["total_ft"],
            "succ_pct":      f"{stats['total_succ']/total_tr*100:.1f}" if total_tr else "0",
            "ff_pct":        f"{stats['total_ff']/total_tr*100:.1f}"   if total_tr else "0",
            "fk_pct":        f"{stats['total_fk']/total_tr*100:.1f}"   if total_tr else "0",
            "ft_pct":        f"{stats['total_ft']/total_tr*100:.1f}"   if total_tr else "0",
            "overall_rate":  f"{stats['total_succ']/total_tr*100:.1f}" if total_tr else "0",
            "use_dist": _dist_rows(tu_vals, [
                ("1-5次",1,6),("6-15次",6,16),("16-30次",16,31),
                ("31-50次",31,51),(">50次",51,10**9)
            ]) if tu_vals else [],
            "rate_sessions_count": len(rate_sessions),
            "rate_dist": _dist_rows(
                [s["tool_success_rate"] for s in rate_sessions],
                [("0-50%",0,50),("50-80%",50,80),("80-95%",80,95),
                 ("95-99%",95,99),("100%",100,101)]
            ) if rate_sessions else [],
        },

        "duration": {
            "single_api_count": sum(1 for s in sessions if s["api_call_count"] == 1),
            "multi_count":      len(multi_timed),
            "has_multi":        bool(multi_timed),
            "avg": fmt_duration(sum(dur_vals)/len(dur_vals)) if dur_vals else "N/A",
            "max": fmt_duration(max(dur_vals))               if dur_vals else "N/A",
            "min": fmt_duration(min(dur_vals))               if dur_vals else "N/A",
            "p50": fmt_duration(pct(dur_vals, 50))           if dur_vals else "N/A",
            "p90": fmt_duration(pct(dur_vals, 90))           if dur_vals else "N/A",
            "dist": _dist_rows(dur_vals, [
                ("<1min",0,60),("1-5min",60,300),("5-15min",300,900),
                ("15-30min",900,1800),(">30min",1800,10**9)
            ]) if dur_vals else [],
        },

        "models": [
            {"name": mdl or "(未知)", "count": cnt, "pct": f"{cnt/total*100:.1f}"}
            for mdl, cnt in sorted(stats["model_dist"].items(), key=lambda x: -x[1])
        ],

        "skills": {
            "total_use": sum(stats["global_skills"].values()),
            "distinct_count": len(stats["global_skills"]),
            "top10": [
                {
                    "name": name,
                    "count": count,
                    "pct": f"{count / sum(stats['global_skills'].values()) * 100:.1f}" if sum(stats["global_skills"].values()) else "0",
                }
                for name, count in stats["global_skills"].most_common(10)
            ],
            "has_skills": bool(stats["global_skills"]),
        },

        "quality": {
            "ok_count":   ok_count,
            "ok_pct":     f"{ok_count/total*100:.1f}",
            "fail_count": fail_count,
            "fail_pct":   f"{fail_count/total*100:.1f}",
            "has_fails":  fail_count > 0,
            "error_dist": [
                {
                    "code":  code,
                    "desc":  QUALITY_ERRORS.get(code, code),
                    "count": cnt,
                    "pct":   f"{cnt/total*100:.1f}",
                }
                for code, cnt in sorted(err_counter.items(), key=lambda x: -x[1])
            ],
        },

        "top_sessions": [
            {
                "session":     s["session"],
                "duration":    fmt_duration(s["duration_s"]),
                "user_turns":  s["user_turns"],
                "tool_use":    s["tool_use_count"],
                "tool_result": s["tool_result_count"],
                "rate":        fmt_rate(s["tool_success_rate"]),
                "api_count":   s["api_call_count"],
                "completed":      s["completed"],
                "completed_note": s.get("completed_note", ""),
            }
            for s in top_raw
        ],

        "tools_top10": [
            {
                "name":    name,
                "calls":   calls,
                "success": stats["global_success"].get(name, 0),
                "fail":    stats["global_fail"].get(name, 0),
                "rate":    f"{stats['global_success'].get(name, 0) / calls * 100:.1f}" if calls else "0",
            }
            for name, calls in stats["global_use"].most_common(10)
        ],
    }


# ---------------------------------------------------------------------------
# Excel 导出
# ---------------------------------------------------------------------------

# openpyxl 非法字符（ASCII 控制字符）过滤
_ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]")
_EXCEL_CELL_MAX_LEN = 32767
_Q1_MAX_LEN = 2000


def _sanitize_cell(val: Any, max_len: int = _EXCEL_CELL_MAX_LEN) -> Any:
    """清理 Excel 非法字符，并将字符串截断到单元格允许的最大长度。"""
    if not isinstance(val, str):
        return val
    cleaned = _ILLEGAL_CHARS_RE.sub("", val)
    if len(cleaned) > max_len:
        if max_len <= 1:
            return cleaned[:max_len]
        return cleaned[: max_len - 1] + "…"
    return cleaned


def _fmt_tool_dict(d: Any) -> str:
    """将工具调用 dict 转为 'tool:count, ...' 字符串，按次数降序。"""
    if not isinstance(d, dict) or not d:
        return ""
    return ", ".join(f"{k}:{v}" for k, v in sorted(d.items(), key=lambda x: -x[1]))


def _fmt_skill_dict(d: Any) -> str:
    """将技能 dict 转为 'skillA, skillB' 字符串，按次数降序。"""
    if not isinstance(d, dict) or not d:
        return ""
    return ", ".join(k for k, _ in sorted(d.items(), key=lambda x: (-x[1], x[0])))


# 列定义：(字段key, 中文表头)
_DETAIL_COLS: List[Tuple[str, str]] = [
    ("q1",                "Q1首问"),
    ("session",           "Session"),
    ("start_time",        "开始时间"),
    ("end_time",          "结束时间"),
    ("duration_s",        "持续时长(s)"),
    ("api_call_count",    "请求次数"),
    ("api_errors",        "API错误次数"),
    ("user_turns",        "用户轮次"),
    ("total_messages",    "消息总数"),
    ("tool_use_count",    "tool_use次数"),
    ("tool_result_count", "tool_result次数"),
    ("tool_success",      "工具成功次数"),
    ("tool_fail_flag",    "失败(is_error标记)"),
    ("tool_fail_keyword", "失败(错误关键字)"),
    ("tool_fail_total",   "失败合计"),
    ("tool_success_rate", "工具成功率(%)"),
    ("model",             "模型"),
    ("tool_use_detail",   "工具调用详情"),
    ("tool_success_detail", "工具成功详情"),
    ("skills_used",       "使用的技能"),
    ("completed",         "任务完成"),
    ("completed_note",    "错误备注"),
]


def _bucket_df(pd, vals: List[float], buckets: List[Tuple], unit: str = ""):
    total = len(vals)
    rows = []
    for label, lo, hi in buckets:
        cnt = sum(1 for v in vals if v is not None and lo <= v < hi)
        rows.append({
            "区间": f"{label}{unit}",
            "数量": cnt,
            "占比(%)": round(cnt / total * 100, 1) if total else 0,
        })
    return pd.DataFrame(rows)


def write_excel(sessions: List[Dict], stats: Dict, path: Path) -> None:
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    turns_vals = stats["turns_vals"]
    msg_vals   = stats["msg_vals"]
    api_vals   = stats["api_vals"]
    tu_vals    = stats["tu_vals"]
    rate_vals  = stats["rate_vals"]
    dur_vals   = stats["dur_vals"]

    rows = [
        {label: (
            _sanitize_cell(s.get("q1"), max_len=_Q1_MAX_LEN) if key == "q1"
            else _sanitize_cell(_fmt_skill_dict(s.get(key))) if key == "skills_used"
            else _sanitize_cell(_fmt_tool_dict(s.get(key))) if isinstance(s.get(key), dict)
            else _sanitize_cell(s.get(key))
        )
         for key, label in _DETAIL_COLS}
        for s in sessions
    ]
    df_detail = pd.DataFrame(rows)

    # ── 构建分布 DataFrames ──────────────────────────────────────────────────
    dist_sections = [
        ("对话轮次分布",
         _bucket_df(pd, turns_vals,
                    [(1,1,2),(2,2,4),(4,4,8),(8,8,16),(16,16,10**9)], "轮")),
        ("消息总数分布",
         _bucket_df(pd, msg_vals,
                    [("1-2",0,3),("3-5",3,6),("6-10",6,11),
                     ("11-20",11,21),("21-50",21,51),(">50",51,10**9)], "条")),
        ("API Call次数分布",
         _bucket_df(pd, api_vals,
                    [(1,1,2),(2,2,4),("4-10",4,11),("11-30",11,31),(">30",31,10**9)], "次")),
        ("tool_use次数分布（有工具session）",
         _bucket_df(pd, tu_vals,
                    [("1-5",1,6),("6-15",6,16),("16-30",16,31),("31-50",31,51),(">50",51,10**9)], "次")),
        ("工具成功率分布",
         _bucket_df(pd, rate_vals,
                    [("0-50%",0,50),("50-80%",50,80),("80-95%",80,95),
                     ("95-99%",95,99),("100%",100,101)])),
        ("耗时分布（多轮session）",
         _bucket_df(pd, dur_vals,
                    [("<1min",0,60),("1-5min",60,300),("5-15min",300,900),
                     ("15-30min",900,1800),(">30min",1800,10**9)])),
    ]

    # ── 写入 Excel ───────────────────────────────────────────────────────────
    with pd.ExcelWriter(path, engine="openpyxl") as ew:
        # Sheet 1: 详情
        df_detail.to_excel(ew, sheet_name="Session详情", index=False)
        ws1 = ew.sheets["Session详情"]

        # 冻结首行首列
        ws1.freeze_panes = "B2"

        # 表头样式
        hdr_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
        hdr_font = Font(bold=True, color="FFFFFF", size=10)
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in ws1[1]:
            cell.fill  = hdr_fill
            cell.font  = hdr_font
            cell.alignment = hdr_align
        ws1.row_dimensions[1].height = 30

        # 成功率列转为小数（Excel百分比格式）
        rate_col_idx = next(
            i for i, (_, lbl) in enumerate(_DETAIL_COLS, 1) if lbl == "工具成功率(%)"
        )
        rate_letter = get_column_letter(rate_col_idx)
        for row in range(2, ws1.max_row + 1):
            cell = ws1[f"{rate_letter}{row}"]
            if cell.value is not None:
                try:
                    cell.value = float(cell.value) / 100
                    cell.number_format = "0.0%"
                except (TypeError, ValueError):
                    pass

        # Q1列：统一顶对齐
        q1_col_idx = next(
            i for i, (_, lbl) in enumerate(_DETAIL_COLS, 1) if lbl == "Q1首问"
        )
        q1_letter = get_column_letter(q1_col_idx)
        for row_idx in range(2, ws1.max_row + 1):
            cell = ws1[f"{q1_letter}{row_idx}"]
            cell.alignment = Alignment(vertical="top")

        # 任务完成列：左对齐
        completed_col_idx = next(
            i for i, (_, lbl) in enumerate(_DETAIL_COLS, 1) if lbl == "任务完成"
        )
        completed_letter = get_column_letter(completed_col_idx)
        for row_idx in range(2, ws1.max_row + 1):
            ws1[f"{completed_letter}{row_idx}"].alignment = Alignment(horizontal="left")

        # 固定列宽，避免遍历整张表做自适应
        detail_widths = {
            "Q1首问": 50,
            "Session": 24,
            "开始时间": 20,
            "结束时间": 20,
            "持续时长(s)": 12,
            "请求次数": 10,
            "API错误次数": 12,
            "用户轮次": 10,
            "消息总数": 10,
            "tool_use次数": 12,
            "tool_result次数": 14,
            "工具成功次数": 12,
            "失败(is_error标记)": 16,
            "失败(错误关键字)": 16,
            "失败合计": 10,
            "工具成功率(%)": 12,
            "模型": 20,
            "工具调用详情": 32,
            "工具成功详情": 32,
            "使用的技能": 24,
            "任务完成": 14,
            "错误备注": 28,
        }
        for col_idx, (_key, label) in enumerate(_DETAIL_COLS, start=1):
            ws1.column_dimensions[get_column_letter(col_idx)].width = detail_widths.get(label, 20)

        # Sheet 2: 分布统计
        ws2 = ew.book.create_sheet("分布统计")
        title_font  = Font(bold=True, size=11, color="1F4E79")
        subhdr_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
        subhdr_font = Font(bold=True, size=10)

        cur_row = 1
        for section_title, df_s in dist_sections:
            # 区间标题
            title_cell = ws2.cell(cur_row, 1, section_title)
            title_cell.font = title_font
            cur_row += 1

            # 表头行
            for ci, col_name in enumerate(df_s.columns, 1):
                hc = ws2.cell(cur_row, ci, col_name)
                hc.fill = subhdr_fill
                hc.font = subhdr_font
                hc.alignment = Alignment(horizontal="center")
            cur_row += 1

            # 数据行
            for _, row in df_s.iterrows():
                for ci, val in enumerate(row, 1):
                    ws2.cell(cur_row, ci, _sanitize_cell(val))
                cur_row += 1

            cur_row += 2  # 空行分隔

        # 列宽
        for col in ws2.columns:
            ws2.column_dimensions[get_column_letter(col[0].column)].width = 32

        # Sheet 3: 技能统计
        ws3 = ew.book.create_sheet("技能统计")
        ws3.cell(1, 1, "技能名称").font = subhdr_font
        ws3.cell(1, 1).fill = subhdr_fill
        ws3.cell(1, 2, "使用次数").font = subhdr_font
        ws3.cell(1, 2).fill = subhdr_fill
        ws3.cell(1, 2).alignment = Alignment(horizontal="center")

        # 统计全局技能使用
        global_skills: Counter = Counter()
        for s in sessions:
            global_skills.update(s.get("skills_used", {}))

        row_idx = 2
        for skill_name, count in sorted(global_skills.items(), key=lambda x: (-x[1], x[0])):
            ws3.cell(row_idx, 1, _sanitize_cell(skill_name))
            ws3.cell(row_idx, 2, count)
            ws3.cell(row_idx, 2).alignment = Alignment(horizontal="center")
            row_idx += 1

        ws3.column_dimensions["A"].width = 32
        ws3.column_dimensions["B"].width = 14


# ---------------------------------------------------------------------------
# 报告渲染（Jinja2）
# ---------------------------------------------------------------------------

def render_report(template_name: str, context: Dict, output_path: Path) -> None:
    """用指定 Jinja2 模板渲染报告并写入文件。模板目录为脚本同级的 templates/。"""
    from jinja2 import Environment, FileSystemLoader
    templates_dir = Path(__file__).parent / "templates"
    env = Environment(
        loader=FileSystemLoader(str(templates_dir)),
        autoescape=False,
        keep_trailing_newline=True,
        trim_blocks=True,
        lstrip_blocks=True,
    )
    output_path.write_text(env.get_template(template_name).render(**context), encoding="utf-8")


def save_analysis_cache(sessions: List[Dict], path: Path) -> None:
    """保存分析中间结果，避免导出失败时重新扫描全部 session。"""
    payload = {
        "version": 1,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "session_count": len(sessions),
        "sessions": sessions,
    }
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def load_analysis_cache(path: Path) -> List[Dict]:
    """加载分析中间结果。"""
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, dict) and isinstance(payload.get("sessions"), list):
        return payload["sessions"]
    if isinstance(payload, list):
        return payload
    raise ValueError(f"缓存格式不正确: {path}")


# ---------------------------------------------------------------------------
# 多目录批量输出
# ---------------------------------------------------------------------------

def _slugify_name(text: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9._-]+", "-", text.strip())
    slug = re.sub(r"-+", "-", slug).strip("-_.")
    return slug or "session"


def _dedupe_names(names: List[str]) -> List[str]:
    counter: Counter = Counter()
    result: List[str] = []
    for name in names:
        counter[name] += 1
        result.append(name if counter[name] == 1 else f"{name}-{counter[name]}")
    return result


def _discover_session_roots(paths: List[str], parent_paths: List[str]) -> List[Path]:
    roots: List[Path] = []
    seen: set[Path] = set()

    for raw in paths:
        path = Path(raw).resolve()
        if not path.is_dir():
            raise FileNotFoundError(f"目录不存在: {path}")
        if path not in seen:
            roots.append(path)
            seen.add(path)

    for raw in parent_paths:
        parent = Path(raw).resolve()
        if not parent.is_dir():
            raise FileNotFoundError(f"父目录不存在: {parent}")
        for child in sorted(p.resolve() for p in parent.iterdir() if p.is_dir()):
            if child not in seen:
                roots.append(child)
                seen.add(child)

    return roots


def _resolve_output_dir(
    session_dir: Path,
    base_output_dir: Optional[Path],
    session_name: Optional[str],
    multi_mode: bool,
) -> Path:
    if base_output_dir is None:
        return session_dir / "stat"
    if multi_mode:
        return base_output_dir / (session_name or _slugify_name(session_dir.name))
    return base_output_dir


def _write_server_session_config(
    output_dir: Path,
    session_sources: List[Dict[str, str]],
    host: str,
    port: int,
    log_level: str,
) -> Path:
    config = {
        "mode": "session",
        "session_sources": session_sources,
        "host": host,
        "port": port,
        "log_level": log_level,
    }
    config_path = output_dir / "server_session.yaml"
    config_path.write_text(
        yaml.safe_dump(config, sort_keys=False, allow_unicode=True),
        encoding="utf-8",
    )
    return config_path


def analyze_session_root(
    session_dir: Path,
    output_dir: Path,
    recompute: bool,
    worker_num: int,
) -> Dict[str, str]:
    output_dir.mkdir(parents=True, exist_ok=True)

    logger.info("扫描目录: %s", session_dir)
    folders = sorted(
        f for f in session_dir.iterdir()
        if f.is_dir() and f.resolve() != output_dir
    )
    logger.info("发现 %d 个 session 文件夹", len(folders))

    analysis_cache_path = output_dir / "session_analysis.json"
    sessions: List[Dict]
    load_or_scan_start = time.perf_counter()
    if analysis_cache_path.exists() and not recompute:
        logger.info("开始加载分析缓存: %s", analysis_cache_path)
        sessions = load_analysis_cache(analysis_cache_path)
        logger.info("复用分析缓存: %s", analysis_cache_path)
        logger.info("缓存中包含 %d 个 session", len(sessions))
    else:
        sessions = []
        logger.info("开始并行解析 session，worker_num=%d", worker_num)
        with ProcessPoolExecutor(max_workers=worker_num) as executor:
            futures = {executor.submit(analyze_session, f): f for f in folders}
            for future in tqdm(as_completed(futures), total=len(futures), desc=f"解析 {session_dir.name}", unit="个"):
                r = future.result()
                if r:
                    sessions.append(r)
        logger.info("成功解析 %d 个 session", len(sessions))
        logger.info("开始写分析缓存: %s", analysis_cache_path)
        save_analysis_cache(sessions, analysis_cache_path)
        logger.info("分析缓存已写入: %s", analysis_cache_path)
    load_or_scan_elapsed = time.perf_counter() - load_or_scan_start

    logger.info("开始聚合统计")
    stats_start = time.perf_counter()
    stats = compute_stats(sessions)
    stats_elapsed = time.perf_counter() - stats_start
    logger.info("开始构建报告上下文")
    context_start = time.perf_counter()
    ctx = build_context(sessions, stats)
    context_elapsed = time.perf_counter() - context_start

    xlsx_path = output_dir / "session_report.xlsx"
    logger.info("开始写 Excel: %s", xlsx_path)
    excel_start = time.perf_counter()
    write_excel(sessions, stats, xlsx_path)
    excel_elapsed = time.perf_counter() - excel_start
    logger.info("Excel 已写入: %s", xlsx_path)

    html_path = output_dir / "session_report.html"
    logger.info("开始渲染 HTML: %s", html_path)
    html_start = time.perf_counter()
    render_report("report.html.j2", ctx, html_path)
    html_elapsed = time.perf_counter() - html_start
    logger.info("HTML 已写入: %s", html_path)

    md_path = output_dir / "session_report.md"
    logger.info("开始渲染 Markdown: %s", md_path)
    md_start = time.perf_counter()
    render_report("report.md.j2", ctx, md_path)
    md_elapsed = time.perf_counter() - md_start
    logger.info("Markdown 已写入: %s", md_path)
    logger.info(
        "[timing] load_or_scan=%s stats=%s context=%s excel=%s html=%s markdown=%s",
        format_stage_seconds(load_or_scan_elapsed),
        format_stage_seconds(stats_elapsed),
        format_stage_seconds(context_elapsed),
        format_stage_seconds(excel_elapsed),
        format_stage_seconds(html_elapsed),
        format_stage_seconds(md_elapsed),
    )
    return {
        "session_dir": str(session_dir),
        "report_dir": str(output_dir),
        "session_count": str(len(sessions)),
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="分析对话日志目录，支持多个 --dir 批量输出报告，并生成 server_session.yaml",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "示例:\n"
            "  python analyze_sessions.py --dir test_session\n"
            "  python analyze_sessions.py --dir test_session --out /tmp/report\n"
            "  python analyze_sessions.py --dir path_a --dir path_b --out output\n"
            "  python analyze_sessions.py --dir-parent ./sessions --out output"
        ),
    )
    parser.add_argument("paths", metavar="DIR", nargs="*",
                        help="待分析目录（兼容旧用法，可传多个）")
    parser.add_argument("--dir", dest="dirs", action="append", default=[], metavar="DIR",
                        help="待分析目录，可重复传入")
    parser.add_argument("--dir-parent", action="append", default=[], metavar="PARENT_DIR",
                        help="父目录，其直接子目录都会作为待分析目录，可重复传入")
    parser.add_argument("--out", "-o", default=None, metavar="OUTPUT_DIR",
                        help="输出目录。单目录模式默认 <DIR>/stat；多目录模式会输出到 <OUTPUT_DIR>/<name>/")
    parser.add_argument("--recompute", action="store_true",
                        help="忽略已有分析缓存，重新扫描全部 session")
    parser.add_argument("--worker-num", type=int, default=(os.cpu_count() or 1),
                        help="并行 worker 数，默认使用 CPU 核心数")
    parser.add_argument("--server-config-name", default="server_session.yaml",
                        help="批量模式生成的 server 配置文件名，默认 server_session.yaml")
    parser.add_argument("--server-host", default="0.0.0.0",
                        help="生成 server 配置时写入的 host，默认 0.0.0.0")
    parser.add_argument("--server-port", type=int, default=8080,
                        help="生成 server 配置时写入的 port，默认 8080")
    parser.add_argument("--server-log-level", default="INFO",
                        help="生成 server 配置时写入的 log_level，默认 INFO")
    args = parser.parse_args()

    input_dirs = list(args.dirs) + list(args.paths)
    parent_dirs = list(args.dir_parent)

    if not input_dirs and not parent_dirs:
        parser.error("至少指定一个 --dir/目录 或 --dir-parent")

    try:
        session_roots = _discover_session_roots(input_dirs, parent_dirs)
    except FileNotFoundError as exc:
        parser.error(str(exc))

    if not session_roots:
        parser.error("未发现可分析的目录")

    multi_mode = len(session_roots) > 1 or bool(parent_dirs)
    base_output_dir = Path(args.out).resolve() if args.out else None
    if multi_mode and base_output_dir is None:
        parser.error("多目录模式必须显式指定 --out，作为统一输出根目录")

    raw_names = [_slugify_name(root.name) for root in session_roots]
    unique_names = _dedupe_names(raw_names)
    worker_num = max(1, args.worker_num)

    source_entries: List[Dict[str, str]] = []
    for session_dir, session_name in zip(session_roots, unique_names):
        output_dir = _resolve_output_dir(session_dir, base_output_dir, session_name, multi_mode)
        logger.info("开始分析 session 根目录: %s -> %s", session_dir, output_dir)
        analyze_session_root(
            session_dir=session_dir,
            output_dir=output_dir,
            recompute=args.recompute,
            worker_num=worker_num,
        )
        source_entries.append({
            "label": session_name,
            "session_dir": str(session_dir),
            "report_dir": str(output_dir),
        })

    config_output_dir = base_output_dir if multi_mode else _resolve_output_dir(session_roots[0], base_output_dir, unique_names[0], False)
    config_output_dir.mkdir(parents=True, exist_ok=True)
    config_path = config_output_dir / args.server_config_name
    config_path.write_text(
        yaml.safe_dump(
            {
                "mode": "session",
                "session_sources": source_entries,
                "host": args.server_host,
                "port": args.server_port,
                "log_level": args.server_log_level,
            },
            sort_keys=False,
            allow_unicode=True,
        ),
        encoding="utf-8",
    )
    logger.info("server 配置已写入: %s", config_path)


if __name__ == "__main__":
    main()

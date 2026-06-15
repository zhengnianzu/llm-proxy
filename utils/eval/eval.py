"""
utils/eval/eval.py — 质检评估

从 chat-log-viewer/analyze_sessions.py 剥离核心逻辑:
  - analyze_best_data(): 单 session 的消息/工具/质量分析
  - compute_stats(): 聚合统计
  - write_excel(): 多 sheet Excel 报告
"""

import json
import logging
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

from jinja2 import Environment, FileSystemLoader

from utils.eval.quality_rules import QualityContext, evaluate_quality, fmt_quality, QUALITY_ERRORS

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# 时间戳解析
# ---------------------------------------------------------------------------

FNAME_TS_RE = re.compile(r"^(\d{4}-\d{2}-\d{2})_(\d{2})-(\d{2})-(\d{2})_\d{3}$")


def _parse_folder_ts(name: str) -> Optional[datetime]:
    m = FNAME_TS_RE.match(name)
    if not m:
        return None
    try:
        return datetime.strptime(
            f"{m.group(1)} {m.group(2)}:{m.group(3)}:{m.group(4)}",
            "%Y-%m-%d %H:%M:%S",
        )
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# 工具错误关键字
# ---------------------------------------------------------------------------

_ERROR_PATTERNS: List[re.Pattern] = [
    re.compile(r"Traceback \(most recent call last\)", re.I),
    re.compile(
        r"\b(SyntaxError|NameError|TypeError|ValueError|AttributeError|"
        r"ImportError|ModuleNotFoundError|RuntimeError|KeyError|IndexError|"
        r"FileNotFoundError|PermissionError|OSError|IOError|ZeroDivisionError|"
        r"RecursionError|MemoryError)\s*:", re.I
    ),
    re.compile(r"permission denied", re.I),
    re.compile(r"operation not permitted", re.I),
    re.compile(r"access denied", re.I),
    re.compile(r"cannot execute", re.I),
    re.compile(r"no such file or directory", re.I),
    re.compile(r"file not found", re.I),
    re.compile(r"\btimed?\s*out\b", re.I),
    re.compile(r"\bkilled\b", re.I),
    re.compile(r"segmentation fault", re.I),
    re.compile(r"command not found", re.I),
    re.compile(r"\berror\b", re.I),
    re.compile(r"\bfailed\b", re.I),
    re.compile(r"\bexception\b", re.I),
    re.compile(r"\bfailure\b", re.I),
]


def _has_error_keywords(text: str) -> bool:
    return any(p.search(text) for p in _ERROR_PATTERNS)


# ---------------------------------------------------------------------------
# 内容工具函数
# ---------------------------------------------------------------------------

def _collect_text(content: Any) -> str:
    if isinstance(content, str):
        return content
    if isinstance(content, list):
        parts = []
        for blk in content:
            if not isinstance(blk, dict):
                continue
            if blk.get("text"):
                parts.append(str(blk["text"]))
            inner = blk.get("content", "")
            if isinstance(inner, str):
                parts.append(inner)
            elif isinstance(inner, list):
                for sub in inner:
                    if isinstance(sub, dict) and sub.get("text"):
                        parts.append(str(sub["text"]))
        return "\n".join(parts)
    return ""


def _iter_blocks(content: Any):
    if isinstance(content, str):
        yield {"type": "text", "text": content}
    elif isinstance(content, list):
        for b in content:
            if isinstance(b, dict):
                yield b


# ---------------------------------------------------------------------------
# 乱码检测
# ---------------------------------------------------------------------------

def _is_garbled(text: str, min_lines: int = 10, max_avg_chars: float = 5.0) -> bool:
    lines = [l for l in text.splitlines() if l.strip()]
    if len(lines) < min_lines:
        return False
    return (sum(len(l) for l in lines) / len(lines)) < max_avg_chars


def _mark_garbled_from_content(content: Any, stats: Dict[str, Any]) -> None:
    text = _collect_text(content)
    if text and _is_garbled(text):
        stats["has_garbled"] = True
    if isinstance(content, list):
        for blk in content:
            if not isinstance(blk, dict):
                continue
            thinking = blk.get("thinking") or blk.get("reasoning_content") or ""
            if isinstance(thinking, str) and _is_garbled(thinking):
                stats["has_garbled"] = True


# ---------------------------------------------------------------------------
# Q1 提取
# ---------------------------------------------------------------------------

def _extract_first_user_text(content: Any) -> str:
    if isinstance(content, str):
        text = content
    elif isinstance(content, list):
        parts = [b.get("text") or b.get("id") or "" for b in content if isinstance(b, dict)]
        text = "\n".join(p for p in parts if p)
    else:
        text = str(content)

    while True:
        prev = text
        text = re.sub(r"^\s*\[[^\]]*\]\s*", "", text)
        text = re.sub(
            r"^Sender\s*(?:\([^)]*\))?:\s*```json\s*\{[\s\S]*?\}\s*```\s*",
            "", text, flags=re.IGNORECASE,
        )
        text = re.sub(r"^Sender\s*(?:\([^)]*\))?:[^\n]*\n?", "", text, flags=re.IGNORECASE)
        if text == prev:
            break
    return text.strip()


# ---------------------------------------------------------------------------
# 技能识别
# ---------------------------------------------------------------------------

_SKILL_PATH_RE = re.compile(r"(?:^|/)\.openclaw/skills/([^/]+)/SKILL\.md")


def _record_skill_use(name: str, blk: dict, stats: Dict[str, Any]) -> None:
    if name != "read":
        return
    inp = blk.get("input") or {}
    if isinstance(inp, str):
        try:
            inp = json.loads(inp)
        except Exception:
            inp = {}
    if not isinstance(inp, dict):
        inp = {}
    file_path = inp.get("file_path") or inp.get("path") or ""
    if "SKILL.md" not in file_path:
        return
    match = _SKILL_PATH_RE.search(file_path)
    if match:
        stats["skills_used"][match.group(1)] += 1


# ---------------------------------------------------------------------------
# 工具统计
# ---------------------------------------------------------------------------

def _record_tool_use(blk: dict, stats: Dict[str, Any]) -> None:
    tool_id = blk.get("id", "")
    name = blk.get("name", "unknown")
    if tool_id:
        stats["id_to_name"][tool_id] = name
    stats["tool_use_detail"][name] += 1
    stats["tool_use_count"] += 1
    _record_skill_use(name, blk, stats)


def _record_openai_tool_calls(msg: dict, stats: Dict[str, Any]) -> None:
    for tc in (msg.get("tool_calls") or []):
        fn = tc.get("function", {})
        tool_id = tc.get("id", "")
        name = fn.get("name", "unknown")
        if tool_id:
            stats["id_to_name"][tool_id] = name
        stats["tool_use_detail"][name] += 1
        stats["tool_use_count"] += 1


def _record_tool_result(blk: dict, stats: Dict[str, Any]) -> None:
    stats["tool_result_count"] += 1
    name = stats["id_to_name"].get(blk.get("tool_use_id", ""), "unknown")
    if blk.get("is_error"):
        stats["tool_fail_flag"] += 1
        stats["tool_fail_detail"][name] += 1
        return
    text = _collect_text(blk.get("content", ""))
    if _has_error_keywords(text):
        stats["tool_fail_keyword"] += 1
        stats["tool_fail_detail"][name] += 1
    else:
        stats["tool_success"] += 1
        stats["tool_success_detail"][name] += 1


def _analyze_user_message(content: Any, stats: Dict[str, Any]) -> None:
    has_tool_result = False
    for blk in _iter_blocks(content):
        blk_type = blk.get("type")
        if blk_type == "tool_result":
            has_tool_result = True
            _record_tool_result(blk, stats)
    if not has_tool_result:
        stats["user_turns"] += 1


def _analyze_assistant_message(msg: dict, stats: Dict[str, Any]) -> None:
    content = msg.get("content", [])
    for blk in _iter_blocks(content):
        if blk.get("type") == "tool_use":
            _record_tool_use(blk, stats)
    _record_openai_tool_calls(msg, stats)


def _analyze_tool_role_message(msg: dict, stats: Dict[str, Any]) -> None:
    stats["tool_result_count"] += 1
    name = stats["id_to_name"].get(msg.get("tool_call_id", ""), "unknown")
    text = _collect_text(msg.get("content", ""))
    if _has_error_keywords(text):
        stats["tool_fail_keyword"] += 1
        stats["tool_fail_detail"][name] += 1
    else:
        stats["tool_success"] += 1
        stats["tool_success_detail"][name] += 1


# ---------------------------------------------------------------------------
# analyze_best_data — 核心单 session 分析
# ---------------------------------------------------------------------------

def analyze_best_data(best_data: dict) -> Dict[str, Any]:
    messages: List[dict] = best_data.get("messages") or []
    resp = best_data.get("response") or {}
    resp_content = resp.get("content")

    if resp_content is None and resp.get("choices"):
        oai_msg = (resp["choices"][0] or {}).get("message") or {}
        text = oai_msg.get("content")
        if text:
            resp_content = [{"type": "text", "text": text}]
        if oai_msg.get("tool_calls"):
            resp_content = resp_content or []
            for tc in oai_msg["tool_calls"]:
                fn = tc.get("function", {})
                try:
                    inp = json.loads(fn.get("arguments", "{}"))
                except (json.JSONDecodeError, TypeError):
                    inp = fn.get("arguments", "")
                resp_content.append({
                    "type": "tool_use",
                    "id": tc.get("id", ""),
                    "name": fn.get("name", "unknown"),
                    "input": inp,
                })

    stats = {
        "q1": "",
        "user_turns": 0,
        "tool_use_count": 0,
        "tool_result_count": 0,
        "tool_success": 0,
        "tool_fail_flag": 0,
        "tool_fail_keyword": 0,
        "tool_use_detail": Counter(),
        "tool_success_detail": Counter(),
        "tool_fail_detail": Counter(),
        "skills_used": Counter(),
        "id_to_name": {},
        "has_garbled": False,
    }

    for msg in messages:
        role = msg.get("role")
        content = msg.get("content", [])

        if role == "user" and not stats["q1"]:
            stats["q1"] = _extract_first_user_text(content)

        if role == "user":
            _analyze_user_message(content, stats)
        elif role == "assistant":
            _analyze_assistant_message(msg, stats)
        elif role == "tool":
            _analyze_tool_role_message(msg, stats)

        _mark_garbled_from_content(content, stats)

    if isinstance(resp_content, list):
        for blk in resp_content:
            if isinstance(blk, dict) and blk.get("type") == "tool_use":
                _record_tool_use(blk, stats)
        _mark_garbled_from_content(resp_content, stats)

    tool_result_count = stats["tool_result_count"]

    quality_ctx = QualityContext(resp=resp, resp_content=resp_content, stats=stats)

    return {
        "q1": stats["q1"],
        "total_messages": len(messages) + (1 if resp_content else 0),
        "user_turns": stats["user_turns"],
        "tool_use_count": stats["tool_use_count"],
        "tool_result_count": tool_result_count,
        "tool_success": stats["tool_success"],
        "tool_fail_flag": stats["tool_fail_flag"],
        "tool_fail_keyword": stats["tool_fail_keyword"],
        "tool_fail_total": stats["tool_fail_flag"] + stats["tool_fail_keyword"],
        "tool_success_rate": (
            round(stats["tool_success"] / tool_result_count * 100, 1) if tool_result_count > 0 else None
        ),
        "tool_use_detail": dict(stats["tool_use_detail"]),
        "tool_success_detail": dict(stats["tool_success_detail"]),
        "tool_fail_detail": dict(stats["tool_fail_detail"]),
        "skills_used": dict(stats["skills_used"]),
        "has_garbled": stats["has_garbled"],
        "quality_errors": evaluate_quality(quality_ctx),
    }


# ---------------------------------------------------------------------------
# analyze_session — 单 session 文件夹分析
# ---------------------------------------------------------------------------

def analyze_session(folder: Path, index_meta: Optional[Dict] = None) -> Optional[Dict]:
    json_files = sorted(folder.glob("*.json"))
    if not json_files:
        return None

    start_ts = _parse_folder_ts(folder.name)
    end_ts = _parse_folder_ts(json_files[-1].stem)
    duration_s: Optional[float] = None
    if start_ts and end_ts and end_ts >= start_ts:
        duration_s = (end_ts - start_ts).total_seconds()

    if index_meta and index_meta.get("trace_list"):
        api_call_count = len(index_meta["trace_list"])
    else:
        api_call_count = len(json_files)
    api_errors = 0

    try:
        best_data = json.loads(json_files[-1].read_text(encoding="utf-8"))
    except Exception:
        return None

    resp = best_data.get("response") or {}
    if isinstance(resp, dict) and isinstance(resp.get("status_code"), int):
        if resp["status_code"] >= 400:
            api_errors = 1

    analyzed = analyze_best_data(best_data)

    return {
        "session": folder.name,
        "start_time": start_ts.strftime("%Y-%m-%d %H:%M:%S") if start_ts else None,
        "end_time": end_ts.strftime("%Y-%m-%d %H:%M:%S") if end_ts else None,
        "duration_s": duration_s,
        "api_call_count": api_call_count,
        "api_errors": api_errors,
        "user_turns": analyzed["user_turns"],
        "total_messages": analyzed["total_messages"],
        "tool_use_count": analyzed["tool_use_count"],
        "tool_result_count": analyzed["tool_result_count"],
        "tool_success": analyzed["tool_success"],
        "tool_fail_flag": analyzed["tool_fail_flag"],
        "tool_fail_keyword": analyzed["tool_fail_keyword"],
        "tool_fail_total": analyzed["tool_fail_total"],
        "tool_success_rate": analyzed["tool_success_rate"],
        "model": best_data.get("model", ""),
        "q1": analyzed["q1"],
        "tool_use_detail": analyzed["tool_use_detail"],
        "tool_success_detail": analyzed["tool_success_detail"],
        "tool_fail_detail": analyzed["tool_fail_detail"],
        "skills_used": analyzed["skills_used"],
        **dict(zip(("completed", "completed_note"), fmt_quality(analyzed["quality_errors"]))),
    }


# ---------------------------------------------------------------------------
# 统计聚合
# ---------------------------------------------------------------------------

def _fmt_duration(s: Optional[float]) -> str:
    if s is None:
        return "N/A"
    if s < 60:
        return f"{s:.0f}s"
    if s < 3600:
        return f"{s / 60:.1f}min"
    return f"{s / 3600:.1f}h"


def _pct(values: List[float], p: int) -> float:
    if not values:
        return 0.0
    sv = sorted(values)
    return sv[min(int(len(sv) * p / 100), len(sv) - 1)]


def compute_stats(sessions: List[Dict]) -> Dict:
    turns_vals = [s["user_turns"] for s in sessions]
    msg_vals = [s["total_messages"] for s in sessions]
    api_vals = [s["api_call_count"] for s in sessions]
    with_tools = [s for s in sessions if s["tool_use_count"] > 0]
    tu_vals = [s["tool_use_count"] for s in with_tools]
    rate_vals = [s["tool_success_rate"] for s in sessions if s["tool_success_rate"] is not None]
    multi_timed = [s for s in sessions if s["api_call_count"] > 1 and s["duration_s"] is not None]
    dur_vals = [s["duration_s"] for s in multi_timed]

    global_use: Counter = Counter()
    global_success: Counter = Counter()
    global_fail: Counter = Counter()
    global_skills: Counter = Counter()
    for s in sessions:
        global_use.update(s.get("tool_use_detail", {}))
        global_success.update(s.get("tool_success_detail", {}))
        global_fail.update(s.get("tool_fail_detail", {}))
        global_skills.update(s.get("skills_used", {}))

    return {
        "total": len(sessions),
        "turns_vals": turns_vals,
        "msg_vals": msg_vals,
        "api_vals": api_vals,
        "with_tools": with_tools,
        "tu_vals": tu_vals,
        "rate_vals": rate_vals,
        "multi_timed": multi_timed,
        "dur_vals": dur_vals,
        "rate_sessions": [s for s in sessions if s["tool_result_count"] > 0],
        "model_dist": Counter(s["model"] for s in sessions),
        "multi_api": sum(1 for v in api_vals if v > 1),
        "api_err": sum(1 for s in sessions if s["api_errors"] > 0),
        "total_tu": sum(s["tool_use_count"] for s in sessions),
        "total_tr": sum(s["tool_result_count"] for s in sessions),
        "total_succ": sum(s["tool_success"] for s in sessions),
        "total_ff": sum(s["tool_fail_flag"] for s in sessions),
        "total_fk": sum(s["tool_fail_keyword"] for s in sessions),
        "total_ft": sum(s["tool_fail_total"] for s in sessions),
        "global_use": global_use,
        "global_success": global_success,
        "global_fail": global_fail,
        "global_skills": global_skills,
    }


# ---------------------------------------------------------------------------
# Excel 导出
# ---------------------------------------------------------------------------

_ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]")
_EXCEL_CELL_MAX_LEN = 32767
_Q1_MAX_LEN = 2000


def _sanitize_cell(val: Any, max_len: int = _EXCEL_CELL_MAX_LEN) -> Any:
    if not isinstance(val, str):
        return val
    cleaned = _ILLEGAL_CHARS_RE.sub("", val)
    if len(cleaned) > max_len:
        if max_len <= 1:
            return cleaned[:max_len]
        return cleaned[:max_len - 1] + "…"
    return cleaned


def _fmt_tool_dict(d: Any) -> str:
    if not isinstance(d, dict) or not d:
        return ""
    return ", ".join(f"{k}:{v}" for k, v in sorted(d.items(), key=lambda x: -x[1]))


def _fmt_skill_dict(d: Any) -> str:
    if not isinstance(d, dict) or not d:
        return ""
    return ", ".join(k for k, _ in sorted(d.items(), key=lambda x: (-x[1], x[0])))


_DETAIL_COLS: List[Tuple[str, str]] = [
    ("q1", "Q1首问"),
    ("session", "Session"),
    ("start_time", "开始时间"),
    ("end_time", "结束时间"),
    ("duration_s", "持续时长(s)"),
    ("api_call_count", "请求次数"),
    ("api_errors", "API错误次数"),
    ("user_turns", "用户轮次"),
    ("total_messages", "消息总数"),
    ("tool_use_count", "tool_use次数"),
    ("tool_result_count", "tool_result次数"),
    ("tool_success", "工具成功次数"),
    ("tool_fail_flag", "失败(is_error标记)"),
    ("tool_fail_keyword", "失败(错误关键字)"),
    ("tool_fail_total", "失败合计"),
    ("tool_success_rate", "工具成功率(%)"),
    ("model", "模型"),
    ("tool_use_detail", "工具调用详情"),
    ("tool_success_detail", "工具成功详情"),
    ("skills_used", "使用的技能"),
    ("completed", "任务完成"),
    ("completed_note", "错误备注"),
]


def _bucket_rows(vals: List[float], buckets: List[Tuple], unit: str = "") -> List[Tuple[str, int, str]]:
    total = len(vals)
    rows = []
    for label, lo, hi in buckets:
        cnt = sum(1 for v in vals if v is not None and lo <= v < hi)
        pct = f"{cnt / total * 100:.1f}" if total else "0.0"
        rows.append((f"{label}{unit}", cnt, pct))
    return rows


def _bucket_df(pd, vals: List[float], buckets: List[Tuple], unit: str = ""):
    total = len(vals)
    rows = []
    for label, lo, hi in buckets:
        cnt = sum(1 for v in vals if v is not None and lo <= v < hi)
        rows.append({
            "区间": f"{label}{unit}",
            "数量": cnt,
            "占比(%)": round(cnt / total * 100, 1) if total else 0,
        })
    return pd.DataFrame(rows)


def _md_table(headers: List[str], rows: List[List[str]]) -> str:
    lines = ["| " + " | ".join(headers) + " |"]
    lines.append("|" + "|".join("---" for _ in headers) + "|")
    for row in rows:
        lines.append("| " + " | ".join(str(c) for c in row) + " |")
    return "\n".join(lines)


def write_markdown(sessions: List[Dict], stats: Dict, path: Path) -> None:
    total = stats["total"]
    total_tu = stats["total_tu"]
    total_tr = stats["total_tr"]
    total_succ = stats["total_succ"]
    succ_rate = f"{total_succ / total_tr * 100:.1f}" if total_tr > 0 else "0.0"

    model_dist = stats["model_dist"]
    model_str = ", ".join(f"{m}: {c}" for m, c in model_dist.most_common())

    lines = [
        "# Session 质检报告",
        "",
        f"> 总 Sessions: {total} | 工具调用: {total_tu} | 成功: {total_succ} ({succ_rate}%)",
        f"> 模型: {model_str} | 多轮: {stats['multi_api']} | API错误: {stats['api_err']}",
        "",
    ]

    # --- Session 详情（精简 9 列）---
    lines.append("## Session 详情")
    lines.append("")
    detail_headers = ["Session", "Q1", "时长(s)", "轮次", "消息", "工具调用", "成功率", "模型", "质检"]
    detail_rows = []
    for s in sessions:
        sess = s.get("session", "")
        if len(sess) > 20:
            sess = sess[:20]
        q1 = (s.get("q1") or "")[:80]
        q1 = q1.replace("|", "\\|").replace("\n", " ")
        dur = s.get("duration_s")
        dur_s = f"{dur:.0f}" if dur is not None else "-"
        rate = s.get("tool_success_rate")
        rate_s = f"{rate:.1f}%" if rate is not None else "-"
        completed = s.get("completed", "0")
        note = s.get("completed_note", "")
        quality = f"{completed}" + (f" {note}" if note and completed != "0" else "")
        detail_rows.append([
            sess, q1, dur_s,
            str(s.get("user_turns", 0)),
            str(s.get("total_messages", 0)),
            str(s.get("tool_use_count", 0)),
            rate_s,
            s.get("model", ""),
            quality,
        ])
    lines.append(_md_table(detail_headers, detail_rows))
    lines.append("")

    # --- 分布统计 ---
    lines.append("## 分布统计")
    lines.append("")

    dist_sections = [
        ("对话轮次分布", stats["turns_vals"],
         [(1, 1, 2), (2, 2, 4), (4, 4, 8), (8, 8, 16), (16, 16, 10**9)], "轮"),
        ("消息总数分布", stats["msg_vals"],
         [("1-2", 0, 3), ("3-5", 3, 6), ("6-10", 6, 11),
          ("11-20", 11, 21), ("21-50", 21, 51), (">50", 51, 10**9)], "条"),
        ("API Call次数分布", stats["api_vals"],
         [(1, 1, 2), (2, 2, 4), ("4-10", 4, 11), ("11-30", 11, 31), (">30", 31, 10**9)], "次"),
        ("tool_use次数分布（有工具session）", stats["tu_vals"],
         [("1-5", 1, 6), ("6-15", 6, 16), ("16-30", 16, 31), ("31-50", 31, 51), (">50", 51, 10**9)], "次"),
        ("工具成功率分布", stats["rate_vals"],
         [("0-50%", 0, 50), ("50-80%", 50, 80), ("80-95%", 80, 95),
          ("95-99%", 95, 99), ("100%", 100, 101)], ""),
        ("耗时分布（多轮session）", stats["dur_vals"],
         [("<1min", 0, 60), ("1-5min", 60, 300), ("5-15min", 300, 900),
          ("15-30min", 900, 1800), (">30min", 1800, 10**9)], ""),
    ]

    for title, vals, buckets, unit in dist_sections:
        lines.append(f"### {title}")
        rows = _bucket_rows(vals, buckets, unit)
        lines.append(_md_table(["区间", "数量", "占比(%)"], [[r[0], str(r[1]), r[2]] for r in rows]))
        lines.append("")

    # --- 技能统计 ---
    global_skills = stats["global_skills"]
    if global_skills:
        lines.append("## 技能统计")
        lines.append("")
        skill_rows = [[name, str(count)] for name, count in sorted(global_skills.items(), key=lambda x: (-x[1], x[0]))]
        lines.append(_md_table(["技能名称", "使用次数"], skill_rows))
        lines.append("")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def write_excel(sessions: List[Dict], stats: Dict, path: Path) -> None:
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    turns_vals = stats["turns_vals"]
    msg_vals = stats["msg_vals"]
    api_vals = stats["api_vals"]
    tu_vals = stats["tu_vals"]
    rate_vals = stats["rate_vals"]
    dur_vals = stats["dur_vals"]

    rows = [
        {label: (
            _sanitize_cell(s.get("q1"), max_len=_Q1_MAX_LEN) if key == "q1"
            else _sanitize_cell(_fmt_skill_dict(s.get(key))) if key == "skills_used"
            else _sanitize_cell(_fmt_tool_dict(s.get(key))) if isinstance(s.get(key), dict)
            else _sanitize_cell(s.get(key))
        )
         for key, label in _DETAIL_COLS}
        for s in sessions
    ]
    df_detail = pd.DataFrame(rows)

    dist_sections = [
        ("对话轮次分布",
         _bucket_df(pd, turns_vals,
                    [(1, 1, 2), (2, 2, 4), (4, 4, 8), (8, 8, 16), (16, 16, 10 ** 9)], "轮")),
        ("消息总数分布",
         _bucket_df(pd, msg_vals,
                    [("1-2", 0, 3), ("3-5", 3, 6), ("6-10", 6, 11),
                     ("11-20", 11, 21), ("21-50", 21, 51), (">50", 51, 10 ** 9)], "条")),
        ("API Call次数分布",
         _bucket_df(pd, api_vals,
                    [(1, 1, 2), (2, 2, 4), ("4-10", 4, 11), ("11-30", 11, 31), (">30", 31, 10 ** 9)], "次")),
        ("tool_use次数分布（有工具session）",
         _bucket_df(pd, tu_vals,
                    [("1-5", 1, 6), ("6-15", 6, 16), ("16-30", 16, 31), ("31-50", 31, 51), (">50", 51, 10 ** 9)],
                    "次")),
        ("工具成功率分布",
         _bucket_df(pd, rate_vals,
                    [("0-50%", 0, 50), ("50-80%", 50, 80), ("80-95%", 80, 95),
                     ("95-99%", 95, 99), ("100%", 100, 101)])),
        ("耗时分布（多轮session）",
         _bucket_df(pd, dur_vals,
                    [("<1min", 0, 60), ("1-5min", 60, 300), ("5-15min", 300, 900),
                     ("15-30min", 900, 1800), (">30min", 1800, 10 ** 9)])),
    ]

    with pd.ExcelWriter(path, engine="openpyxl") as ew:
        df_detail.to_excel(ew, sheet_name="Session详情", index=False)
        ws1 = ew.sheets["Session详情"]
        ws1.freeze_panes = "B2"

        hdr_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
        hdr_font = Font(bold=True, color="FFFFFF", size=10)
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in ws1[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = hdr_align
        ws1.row_dimensions[1].height = 30

        rate_col_idx = next(
            i for i, (_, lbl) in enumerate(_DETAIL_COLS, 1) if lbl == "工具成功率(%)"
        )
        rate_letter = get_column_letter(rate_col_idx)
        for row in range(2, ws1.max_row + 1):
            cell = ws1[f"{rate_letter}{row}"]
            if cell.value is not None:
                try:
                    cell.value = float(cell.value) / 100
                    cell.number_format = "0.0%"
                except (TypeError, ValueError):
                    pass

        q1_col_idx = next(
            i for i, (_, lbl) in enumerate(_DETAIL_COLS, 1) if lbl == "Q1首问"
        )
        q1_letter = get_column_letter(q1_col_idx)
        for row_idx in range(2, ws1.max_row + 1):
            ws1[f"{q1_letter}{row_idx}"].alignment = Alignment(vertical="top")

        completed_col_idx = next(
            i for i, (_, lbl) in enumerate(_DETAIL_COLS, 1) if lbl == "任务完成"
        )
        completed_letter = get_column_letter(completed_col_idx)
        for row_idx in range(2, ws1.max_row + 1):
            ws1[f"{completed_letter}{row_idx}"].alignment = Alignment(horizontal="left")

        detail_widths = {
            "Q1首问": 50, "Session": 24, "开始时间": 20, "结束时间": 20,
            "持续时长(s)": 12, "请求次数": 10, "API错误次数": 12,
            "用户轮次": 10, "消息总数": 10, "tool_use次数": 12,
            "tool_result次数": 14, "工具成功次数": 12, "失败(is_error标记)": 16,
            "失败(错误关键字)": 16, "失败合计": 10, "工具成功率(%)": 12,
            "模型": 20, "工具调用详情": 32, "工具成功详情": 32,
            "使用的技能": 24, "任务完成": 14, "错误备注": 28,
        }
        for col_idx, (_key, label) in enumerate(_DETAIL_COLS, start=1):
            ws1.column_dimensions[get_column_letter(col_idx)].width = detail_widths.get(label, 20)

        # Sheet 2: 分布统计
        ws2 = ew.book.create_sheet("分布统计")
        title_font = Font(bold=True, size=11, color="1F4E79")
        subhdr_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
        subhdr_font = Font(bold=True, size=10)

        cur_row = 1
        for section_title, df_s in dist_sections:
            title_cell = ws2.cell(cur_row, 1, section_title)
            title_cell.font = title_font
            cur_row += 1
            for ci, col_name in enumerate(df_s.columns, 1):
                hc = ws2.cell(cur_row, ci, col_name)
                hc.fill = subhdr_fill
                hc.font = subhdr_font
                hc.alignment = Alignment(horizontal="center")
            cur_row += 1
            for _, row in df_s.iterrows():
                for ci, val in enumerate(row, 1):
                    ws2.cell(cur_row, ci, _sanitize_cell(val))
                cur_row += 1
            cur_row += 2

        for col in ws2.columns:
            ws2.column_dimensions[get_column_letter(col[0].column)].width = 32

        # Sheet 3: 技能统计
        ws3 = ew.book.create_sheet("技能统计")
        ws3.cell(1, 1, "技能名称").font = subhdr_font
        ws3.cell(1, 1).fill = subhdr_fill
        ws3.cell(1, 2, "使用次数").font = subhdr_font
        ws3.cell(1, 2).fill = subhdr_fill
        ws3.cell(1, 2).alignment = Alignment(horizontal="center")

        global_skills: Counter = Counter()
        for s in sessions:
            global_skills.update(s.get("skills_used", {}))

        row_idx = 2
        for skill_name, count in sorted(global_skills.items(), key=lambda x: (-x[1], x[0])):
            ws3.cell(row_idx, 1, _sanitize_cell(skill_name))
            ws3.cell(row_idx, 2, count)
            ws3.cell(row_idx, 2).alignment = Alignment(horizontal="center")
            row_idx += 1

        ws3.column_dimensions["A"].width = 32
        ws3.column_dimensions["B"].width = 14


# ---------------------------------------------------------------------------
# HTML 报告 — build_context + render (移植自 analyze_sessions.py)
# ---------------------------------------------------------------------------

def _fmt_rate(r: Optional[float]) -> str:
    return f"{r:.1f}%" if r is not None else "-"


def _dist_rows(vals: List, buckets: List[Tuple], unit: str = "",
               session_items: Optional[List[Dict]] = None) -> List[Dict]:
    total = len(vals)
    bucket_sessions: List[List[Dict]] = [[] for _ in buckets]
    counts = []
    for bi, (_, lo, hi) in enumerate(buckets):
        cnt = 0
        for vi, v in enumerate(vals):
            if v is not None and lo <= v < hi:
                cnt += 1
                if session_items and vi < len(session_items):
                    bucket_sessions[bi].append(session_items[vi])
        counts.append(cnt)
    return [
        {
            "label": f"{label}{unit}",
            "count": cnt,
            "pct": round(cnt / total * 100, 1) if total else 0.0,
            "sessions": sids,
        }
        for (label, lo, hi), cnt, sids in zip(buckets, counts, bucket_sessions)
    ]


def _extract_error_codes(completed) -> List[str]:
    if completed == 0:
        return []
    codes_part = str(completed).split(" ")[0]
    return [c.strip() for c in codes_part.split(",") if c.strip()]


def build_context(sessions: List[Dict], stats: Dict, top_n: int = 10,
                  key_name: str = "", obs_path: str = "") -> Dict:
    total = stats["total"]
    turns_vals = stats["turns_vals"]
    msg_vals = stats["msg_vals"]
    api_vals = stats["api_vals"]
    with_tools = stats["with_tools"]
    tu_vals = stats["tu_vals"]
    rate_sessions = stats["rate_sessions"]
    multi_timed = stats["multi_timed"]
    dur_vals = stats["dur_vals"]
    total_tr = stats["total_tr"]
    single_count = sum(1 for v in api_vals if v == 1)

    def _si(s):
        return {"name": s["session"], "file": s.get("latest_file", ""), "log_dir": s.get("log_dir", "")}

    all_items = [_si(s) for s in sessions]
    wt_items = [_si(s) for s in with_tools]
    rate_items = [_si(s) for s in rate_sessions]
    mt_items = [_si(s) for s in multi_timed]

    ok_count = sum(1 for s in sessions if s["completed"] == 0)
    fail_count = total - ok_count
    err_counter: Counter = Counter()
    err_sessions: Dict[str, List[Dict]] = {}
    for s in sessions:
        for code in _extract_error_codes(s["completed"]):
            err_counter[code] += 1
            err_sessions.setdefault(code, []).append(_si(s))

    top_raw = sorted(
        [s for s in sessions if s.get("duration_s")],
        key=lambda x: -x["duration_s"],
    )[:top_n]

    return {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "total": total,
        "top_n": top_n,
        "key_name": key_name,
        "obs_path": obs_path,
        "turns": {
            "avg": f"{sum(turns_vals)/total:.2f}",
            "dist": _dist_rows(turns_vals, [
                ("1", 1, 2), ("2-3", 2, 4), ("4-7", 4, 8), ("8-15", 8, 16), (">15", 16, 10**9)
            ], "轮", session_items=all_items),
        },
        "messages": {
            "avg": f"{sum(msg_vals)/total:.2f}",
            "dist": _dist_rows(msg_vals, [
                ("1-2", 0, 3), ("3-5", 3, 6), ("6-10", 6, 11),
                ("11-20", 11, 21), ("21-50", 21, 51), (">50", 51, 10**9)
            ], "条", session_items=all_items),
        },
        "api": {
            "avg": f"{sum(api_vals)/total:.2f}",
            "single_count": single_count,
            "single_pct": f"{single_count/total*100:.1f}",
            "multi_count": stats["multi_api"],
            "multi_pct": f"{stats['multi_api']/total*100:.1f}",
            "err_count": stats["api_err"],
            "err_pct": f"{stats['api_err']/total*100:.1f}",
            "dist": _dist_rows(api_vals, [
                ("1次", 1, 2), ("2-3次", 2, 4), ("4-10次", 4, 11), ("11-30次", 11, 31), (">30次", 31, 10**9)
            ], session_items=all_items),
        },
        "tools": {
            "with_count": len(with_tools),
            "with_pct": f"{len(with_tools)/total*100:.1f}",
            "without_count": total - len(with_tools),
            "without_pct": f"{(total-len(with_tools))/total*100:.1f}",
            "total_use": stats["total_tu"],
            "total_result": total_tr,
            "has_sessions": bool(with_tools),
            "avg_use": f"{sum(tu_vals)/len(tu_vals):.1f}" if tu_vals else "0",
            "has_results": total_tr > 0,
            "total_succ": stats["total_succ"],
            "total_ff": stats["total_ff"],
            "total_fk": stats["total_fk"],
            "total_ft": stats["total_ft"],
            "succ_pct": f"{stats['total_succ']/total_tr*100:.1f}" if total_tr else "0",
            "ff_pct": f"{stats['total_ff']/total_tr*100:.1f}" if total_tr else "0",
            "fk_pct": f"{stats['total_fk']/total_tr*100:.1f}" if total_tr else "0",
            "ft_pct": f"{stats['total_ft']/total_tr*100:.1f}" if total_tr else "0",
            "overall_rate": f"{stats['total_succ']/total_tr*100:.1f}" if total_tr else "0",
            "use_dist": _dist_rows(tu_vals, [
                ("1-5次", 1, 6), ("6-15次", 6, 16), ("16-30次", 16, 31),
                ("31-50次", 31, 51), (">50次", 51, 10**9)
            ], session_items=wt_items) if tu_vals else [],
            "rate_sessions_count": len(rate_sessions),
            "rate_dist": _dist_rows(
                [s["tool_success_rate"] for s in rate_sessions],
                [("0-50%", 0, 50), ("50-80%", 50, 80), ("80-95%", 80, 95),
                 ("95-99%", 95, 99), ("100%", 100, 101)],
                session_items=rate_items,
            ) if rate_sessions else [],
        },
        "duration": {
            "single_api_count": sum(1 for s in sessions if s["api_call_count"] == 1),
            "multi_count": len(multi_timed),
            "has_multi": bool(multi_timed),
            "avg": _fmt_duration(sum(dur_vals)/len(dur_vals)) if dur_vals else "N/A",
            "max": _fmt_duration(max(dur_vals)) if dur_vals else "N/A",
            "dist": _dist_rows(dur_vals, [
                ("<1min", 0, 60), ("1-5min", 60, 300), ("5-15min", 300, 900),
                ("15-30min", 900, 1800), (">30min", 1800, 10**9)
            ], session_items=mt_items) if dur_vals else [],
        },
        "models": [
            {"name": mdl or "(未知)", "count": cnt, "pct": f"{cnt/total*100:.1f}"}
            for mdl, cnt in sorted(stats["model_dist"].items(), key=lambda x: -x[1])
        ],
        "skills": {
            "total_use": sum(stats["global_skills"].values()),
            "distinct_count": len(stats["global_skills"]),
            "top10": [
                {
                    "name": name,
                    "count": count,
                    "pct": f"{count / sum(stats['global_skills'].values()) * 100:.1f}" if sum(stats["global_skills"].values()) else "0",
                }
                for name, count in stats["global_skills"].most_common(10)
            ],
            "has_skills": bool(stats["global_skills"]),
        },
        "quality": {
            "ok_count": ok_count,
            "ok_pct": f"{ok_count/total*100:.1f}",
            "fail_count": fail_count,
            "fail_pct": f"{fail_count/total*100:.1f}",
            "has_fails": fail_count > 0,
            "error_dist": [
                {
                    "code": code,
                    "desc": QUALITY_ERRORS.get(code, code),
                    "count": cnt,
                    "pct": f"{cnt/total*100:.1f}",
                    "sessions": err_sessions.get(code, []),
                }
                for code, cnt in sorted(err_counter.items(), key=lambda x: -x[1])
            ],
        },
        "top_sessions": [
            {
                "session": s["session"],
                "duration": _fmt_duration(s["duration_s"]),
                "user_turns": s["user_turns"],
                "tool_use": s["tool_use_count"],
                "tool_result": s["tool_result_count"],
                "rate": _fmt_rate(s["tool_success_rate"]),
                "api_count": s["api_call_count"],
                "completed": s["completed"],
                "completed_note": s.get("completed_note", ""),
            }
            for s in top_raw
        ],
        "tools_top10": [
            {
                "name": name,
                "calls": calls,
                "success": stats["global_success"].get(name, 0),
                "fail": stats["global_fail"].get(name, 0),
                "rate": f"{stats['global_success'].get(name, 0) / calls * 100:.1f}" if calls else "0",
            }
            for name, calls in stats["global_use"].most_common(10)
        ],
    }


_TEMPLATES_DIR = Path(__file__).parent / "templates"


def render_html_report(sessions: List[Dict], stats: Dict, output_path: Path,
                       key_name: str = "", obs_path: str = "") -> None:
    ctx = build_context(sessions, stats, key_name=key_name, obs_path=obs_path)
    env = Environment(
        loader=FileSystemLoader(str(_TEMPLATES_DIR)),
        autoescape=False,
        keep_trailing_newline=True,
        trim_blocks=True,
        lstrip_blocks=True,
    )
    output_path.write_text(env.get_template("report.html.j2").render(**ctx), encoding="utf-8")


def render_html_report_string(sessions: List[Dict], stats: Dict,
                              key_name: str = "", obs_path: str = "") -> str:
    ctx = build_context(sessions, stats, key_name=key_name, obs_path=obs_path)
    env = Environment(
        loader=FileSystemLoader(str(_TEMPLATES_DIR)),
        autoescape=False,
        keep_trailing_newline=True,
        trim_blocks=True,
        lstrip_blocks=True,
    )
    return env.get_template("report.html.j2").render(**ctx)


def save_analysis_json(sessions: List[Dict], path: Path) -> None:
    payload = {
        "version": 1,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "session_count": len(sessions),
        "sessions": sessions,
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def load_analysis_json(path_or_content) -> List[Dict]:
    if isinstance(path_or_content, Path):
        payload = json.loads(path_or_content.read_text(encoding="utf-8"))
    else:
        payload = json.loads(path_or_content)
    if isinstance(payload, dict) and isinstance(payload.get("sessions"), list):
        return payload["sessions"]
    if isinstance(payload, list):
        return payload
    raise ValueError("Invalid analysis JSON format")


# ---------------------------------------------------------------------------
# 主入口 — evaluate_sessions
# ---------------------------------------------------------------------------

def evaluate_sessions(
    sessions: List[Dict],
    report_dir: str,
    progress_cb: Optional[Callable[[str], None]] = None,
    key_name: str = "",
    obs_path: str = "",
) -> dict:
    """
    接收已分析好的 session 结果列表，聚合统计并生成报告。

    输出: session_report.md + session_report.html + session_analysis.json
    """
    _log = progress_cb or (lambda msg: None)
    out = Path(report_dir)
    out.mkdir(parents=True, exist_ok=True)

    if not sessions:
        _log("无 session 数据")
        return {"total_sessions": 0, "report_path": "", "html_report_path": "",
                "analysis_json_path": "", "stats_summary": {}}

    _log(f"生成报告: {len(sessions)} sessions...")
    stats = compute_stats(sessions)

    xlsx_path = out / "session_report.xlsx"
    write_excel(sessions, stats, xlsx_path)
    _log("session_report.xlsx 已生成")

    report_path = out / "session_report.md"
    write_markdown(sessions, stats, report_path)
    _log("session_report.md 已生成")

    html_path = out / "session_report.html"
    render_html_report(sessions, stats, html_path, key_name=key_name, obs_path=obs_path)
    _log("session_report.html 已生成")

    json_path = out / "session_analysis.json"
    save_analysis_json(sessions, json_path)
    _log("session_analysis.json 已保存")

    return {
        "total_sessions": len(sessions),
        "report_path": str(report_path),
        "html_report_path": str(html_path),
        "analysis_json_path": str(json_path),
        "stats_summary": {
            "total_tool_use": stats["total_tu"],
            "total_tool_result": stats["total_tr"],
            "total_success": stats["total_succ"],
            "total_fail": stats["total_ft"],
            "overall_success_rate": round(stats["total_succ"] / stats["total_tr"] * 100, 1) if stats["total_tr"] > 0 else 0,
        },
    }
